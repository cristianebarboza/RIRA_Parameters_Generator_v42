import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import json
import asyncio
import websockets
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# RIRA Parameters Generator - GAP Data Non-Seedcare (v4.2)
# Syngenta P&S - DPI
# ==============================================================================
# Business rules are FIXED. Users cannot modify rules.
# Users SELECT the export country to determine the applicable RIL.
# ==============================================================================

# -- SYNGENTA COLORS -----------------------------------------------------------
SYNGENTA_GREEN      = "#00A651"
SYNGENTA_DARK_GREEN = "#007A3D"
SYNGENTA_LIGHT_GRAY = "#F5F5F5"
SYNGENTA_GRAY       = "#6D6E71"
SYNGENTA_RED        = "#D32F2F"
SYNGENTA_AMBER      = "#F9A825"

# -- PAGE CONFIG ---------------------------------------------------------------
st.set_page_config(
    page_title="RIRA Parameters Generator | Syngenta",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -- PASSWORD CONFIG
password = st.text_input("Enter access code:", type="password")
if password != st.secrets.get("app-password", ""):
    st.stop()

# -- CSS -----------------------------------------------------------------------
CSS = f"""
<style>
  .main {{ background-color: {SYNGENTA_LIGHT_GRAY}; }}
  .header-bar {{
    background: linear-gradient(90deg, {SYNGENTA_DARK_GREEN}, {SYNGENTA_GREEN});
    padding: 18px 32px; border-radius: 8px; margin-bottom: 24px;
  }}
  .header-bar h1 {{ color:white; font-size:24px; font-weight:700; margin:0; font-family:Arial; }}
  .header-bar p {{ color:rgba(255,255,255,0.85); font-size:13px; margin:4px 0 0 0; font-family:Arial; }}
  .section-title {{
    font-size:15px; font-weight:700; color:{SYNGENTA_DARK_GREEN};
    font-family:Arial; padding:8px 0 4px 0;
    border-bottom:2px solid {SYNGENTA_GREEN}; margin-bottom:12px;
  }}
  .metric-card {{
    background:white; border-radius:8px; padding:18px;
    border-left:5px solid {SYNGENTA_GREEN};
    box-shadow:0 2px 6px rgba(0,0,0,0.07); text-align:center;
  }}
  .metric-card.warn {{ border-left-color:{SYNGENTA_AMBER}; }}
  .metric-card.fail {{ border-left-color:{SYNGENTA_RED}; }}
  .metric-value {{ font-size:28px; font-weight:700; color:{SYNGENTA_DARK_GREEN}; font-family:Arial; }}
  .metric-value.warn {{ color:{SYNGENTA_AMBER}; }}
  .metric-value.fail {{ color:{SYNGENTA_RED}; }}
  .metric-label {{ font-size:11px; color:{SYNGENTA_GRAY}; font-family:Arial; margin-top:4px; }}
  .stButton > button {{
    background-color:{SYNGENTA_GREEN}; color:white;
    border:none; border-radius:6px; font-weight:600;
  }}
  .stButton > button:hover {{ background-color:{SYNGENTA_DARK_GREEN}; }}
  .rule-box {{
    background: white; border-radius: 6px; padding: 12px 16px;
    border-left: 4px solid {SYNGENTA_GREEN}; margin-bottom: 8px;
    font-family: Arial; font-size: 12px;
  }}
  .rule-box.filter {{ border-left-color: {SYNGENTA_AMBER}; }}
  .ril-highlight {{
    background: #E8F5E9; border: 2px solid {SYNGENTA_GREEN};
    border-radius: 8px; padding: 16px; margin: 12px 0;
    font-family: Arial;
  }}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# -- HEADER --------------------------------------------------------------------
st.markdown("""
<div class="header-bar">
  <h1>RIRA Parameters Generator - GAP Non-Seedcare</h1>
  <p>Automated Risk Assessment Parameters | Fixed Business Rules</p>
</div>
""", unsafe_allow_html=True)


# ==============================================================================
#  QLIK CLOUD CONNECTION (Real-time data fetch)
# ==============================================================================

QLIK_TENANT = "qs-syngenta.eu.qlikcloud.com"
QLIK_APP_ID = "d126bbb3-7daa-4f0e-8e73-245e8ea672f2"
QLIK_CLIENT_ID = st.secrets.get("qlik-client-id", "")
QLIK_CLIENT_SECRET = st.secrets.get("qlik-client-secret", "")


def get_qlik_token():
    """Exchange client credentials for OAuth access token."""
    if not QLIK_CLIENT_ID or not QLIK_CLIENT_SECRET:
        return None
    try:
        resp = requests.post(
            f"https://{QLIK_TENANT}/oauth/token",
            data={"grant_type": "client_credentials",
                  "client_id": QLIK_CLIENT_ID,
                  "client_secret": QLIK_CLIENT_SECRET},
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        if resp.status_code == 200:
            return resp.json().get("access_token")
        else:
            st.error(f"Qlik auth failed: {resp.status_code}")
            return None
    except Exception as e:
        st.error(f"Qlik auth error: {e}")
        return None

def trigger_qlik_reload(token):
    """Trigger app reload via Qlik Cloud REST API. Returns True if successful."""
    try:
        resp = requests.post(
            f"https://{QLIK_TENANT}/api/v1/reloads",
            json={"appId": QLIK_APP_ID},
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            },
            timeout=15
        )
        if resp.status_code in (200, 201):
            reload_id = resp.json().get("id")
            # Poll until complete
            for _ in range(60):  # max 5 minutes (5s intervals)
                import time
                time.sleep(5)
                status_resp = requests.get(
                    f"https://{QLIK_TENANT}/api/v1/reloads/{reload_id}",
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=10
                )
                if status_resp.status_code == 200:
                    status = status_resp.json().get("status")
                    if status == "SUCCEEDED":
                        return True
                    elif status in ("FAILED", "CANCELED"):
                        return False
            return False
        else:
            return False
    except Exception:
        return False

def get_qlik_token():
    """Exchange client credentials for OAuth access token."""
    if not QLIK_CLIENT_ID or not QLIK_CLIENT_SECRET:
        return None
    try:
        resp = requests.post(
            f"https://{QLIK_TENANT}/oauth/token",
            data={"grant_type": "client_credentials",
                  "client_id": QLIK_CLIENT_ID,
                  "client_secret": QLIK_CLIENT_SECRET},
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        if resp.status_code == 200:
            return resp.json().get("access_token")
        else:
            st.error(f"Qlik auth failed: {resp.status_code}")
            return None
    except Exception as e:
        st.error(f"Qlik auth error: {e}")
        return None

def trigger_qlik_reload(token):                          # ← ADD HERE
    """Trigger app reload via Qlik Cloud REST API."""
    try:
        resp = requests.post(
            f"https://{QLIK_TENANT}/api/v1/reloads",
            json={"appId": QLIK_APP_ID},
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            },
            timeout=15
        )
        if resp.status_code in (200, 201):
            reload_id = resp.json().get("id")
            import time
            for _ in range(60):
                time.sleep(5)
                status_resp = requests.get(
                    f"https://{QLIK_TENANT}/api/v1/reloads/{reload_id}",
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=10
                )
                if status_resp.status_code == 200:
                    status = status_resp.json().get("status")
                    if status == "SUCCEEDED":
                        return True
                    elif status in ("FAILED", "CANCELED"):
                        return False
            return False
        else:
            return False
    except Exception:
        return False


async def fetch_qlik_data(token, progress_bar):
    ...  # existing code



async def fetch_qlik_data(token, progress_bar):
    """Connect to Qlik via WebSocket and fetch GAP data using known field names."""
    uri = f"wss://{QLIK_TENANT}/app/{QLIK_APP_ID}"
    headers = {"Authorization": f"Bearer {token}"}

    # Known fields from REST API metadata
    fields = [
        "GAP Variant", "Country", "GAP Lifecycle State", "GAP UI State",
        "Target Crop", "Target Crop Code",
        "Max", "UoM",
        "Max. App. Number / Season", "Application Method",
        "GAP Remarks", "GAP Usage Information Type",
        "Seed Care / Non-Seed Care"
    ]
    n_cols = len(fields)

    try:
        async with websockets.connect(uri, extra_headers=headers, open_timeout=30) as ws:
            debug_msgs = []

            # 1. Open doc
            await ws.send(json.dumps({
                "jsonrpc": "2.0", "id": 1, "handle": -1,
                "method": "OpenDoc", "params": [QLIK_APP_ID]
            }))

            dh = None
            for _ in range(20):
                raw = await asyncio.wait_for(ws.recv(), timeout=10)
                msg = json.loads(raw)
                debug_msgs.append(msg)
                if msg.get("id") == 1:
                    if "error" in msg:
                        return None, f"OpenDoc error: {json.dumps(msg['error'])[:300]}"
                    if "result" in msg:
                        qreturn = msg["result"].get("qReturn", {})
                        dh = qreturn.get("qHandle")
                        break
                    else:
                        return None, f"OpenDoc response has id=1 but no result/error: {json.dumps(msg)[:500]}"

            if dh is None:
                summary = [{"id": m.get("id"), "method": m.get("method"),
                           "has_result": "result" in m, "has_error": "error" in m,
                           "keys": list(m.keys())} for m in debug_msgs[:5]]
                return None, f"Could not open app. Messages received: {json.dumps(summary)[:500]}"

            progress_bar.progress(0.03, text="App opened. Applying filters...")

            # ==================================================================
            # 1b. Check last reload time
            # ==================================================================
            await ws.send(json.dumps({
                "jsonrpc": "2.0", "id": 100, "handle": dh,
                "method": "GetAppLayout", "params": []
            }))

            last_reload_time = None
            for _ in range(20):
                raw = await asyncio.wait_for(ws.recv(), timeout=10)
                msg = json.loads(raw)
                if msg.get("id") == 100:
                    if "result" in msg:
                        layout = msg["result"].get("qLayout", {})
                        last_reload_time = layout.get("qLastReloadTime", "Unknown")
                    break

            progress_bar.progress(0.04, text=f"Last data reload: {last_reload_time}")


            # ============================================================
            # 2. Apply filter: "Seed Care / Non-Seed Care" = "Non-Seed Care"
            # ============================================================
            await ws.send(json.dumps({
                "jsonrpc": "2.0", "id": 2, "handle": dh,
                "method": "GetField", "params": [{"qFieldName": "Seed Care / Non-Seed Care"}]
            }))
            field_handle_sc = None
            for _ in range(20):
                raw = await asyncio.wait_for(ws.recv(), timeout=10)
                msg = json.loads(raw)
                if msg.get("id") == 2:
                    if "result" in msg:
                        field_handle_sc = msg["result"]["qReturn"]["qHandle"]
                    break

            if field_handle_sc is not None:
                await ws.send(json.dumps({
                    "jsonrpc": "2.0", "id": 5, "handle": field_handle_sc,
                    "method": "Select",
                    "params": [{"qMatch": "Non-Seed Care", "qSoftLock": False}]
                }))
                for _ in range(20):
                    raw = await asyncio.wait_for(ws.recv(), timeout=10)
                    msg = json.loads(raw)
                    if msg.get("id") == 5:
                        break
                progress_bar.progress(0.05, text="Non-Seedcare filter applied. Applying Approved filter...")
            else:
                progress_bar.progress(0.05, text="⚠️ Seed Care field not found. Applying Approved filter...")

            # ==================================================================
            # 3. Apply filter: "Approved" on lifecycle field
            #    Try "GAP Lifecycle State" first, fall back to "GAP UI State"
            #    Uses SelectValues for reliable exact-match selection
            # ==================================================================
            lifecycle_fields = ["GAP Lifecycle State", "GAP UI State"]
            approved_filter_applied = False

            for lf_idx, lf_name in enumerate(lifecycle_fields):
                lf_msg_id = 6 + (lf_idx * 10)

                # Get field handle
                await ws.send(json.dumps({
                    "jsonrpc": "2.0", "id": lf_msg_id, "handle": dh,
                    "method": "GetField", "params": [{"qFieldName": lf_name}]
                }))

                field_handle_lf = None
                for _ in range(20):
                    raw = await asyncio.wait_for(ws.recv(), timeout=10)
                    msg = json.loads(raw)
                    if msg.get("id") == lf_msg_id:
                        if "result" in msg:
                            field_handle_lf = msg["result"]["qReturn"]["qHandle"]
                        break

                if field_handle_lf is None:
                    continue  # Field doesn't exist, try next

                # Use SelectValues (more reliable than Select with qMatch)
                await ws.send(json.dumps({
                    "jsonrpc": "2.0", "id": lf_msg_id + 1, "handle": field_handle_lf,
                    "method": "SelectValues",
                    "params": [
                        [{"qText": "Approved", "qIsNumeric": False}],
                        False,
                        False
                    ]
                }))

                select_success = False
                for _ in range(20):
                    raw = await asyncio.wait_for(ws.recv(), timeout=10)
                    msg = json.loads(raw)
                    if msg.get("id") == lf_msg_id + 1:
                        if "result" in msg:
                            select_success = msg["result"].get("qReturn", False)
                        break

                if select_success:
                    approved_filter_applied = True
                    progress_bar.progress(0.08,
                        text=f"Approved filter applied on '{lf_name}'. Creating data request...")
                    break
                else:
                    # Exact match failed - try wildcard
                    await ws.send(json.dumps({
                        "jsonrpc": "2.0", "id": lf_msg_id + 2, "handle": field_handle_lf,
                        "method": "Clear", "params": []
                    }))
                    for _ in range(20):
                        raw = await asyncio.wait_for(ws.recv(), timeout=10)
                        msg = json.loads(raw)
                        if msg.get("id") == lf_msg_id + 2:
                            break

                    await ws.send(json.dumps({
                        "jsonrpc": "2.0", "id": lf_msg_id + 3, "handle": field_handle_lf,
                        "method": "Select",
                        "params": [{"qMatch": "Approved*", "qSoftLock": False}]
                    }))
                    for _ in range(20):
                        raw = await asyncio.wait_for(ws.recv(), timeout=10)
                        msg = json.loads(raw)
                        if msg.get("id") == lf_msg_id + 3:
                            if "result" in msg:
                                select_success = msg["result"].get("qReturn", False)
                            break

                    if select_success:
                        approved_filter_applied = True
                        progress_bar.progress(0.08,
                            text=f"Approved filter applied on '{lf_name}' (wildcard). Creating data request...")
                        break

            if not approved_filter_applied:
                progress_bar.progress(0.08,
                    text="⚠️ Could not apply Approved filter. Data may include non-approved records.")
                                                                     

            # ============================================================
            # 4. Create HyperCube (now only fetches Non-Seedcare + Approved)
            # ============================================================
            dims = [{"qDef": {"qFieldDefs": [f], "qFieldLabels": [f]}} for f in fields]
            await ws.send(json.dumps({
                "jsonrpc": "2.0", "id": 3, "handle": dh,
                "method": "CreateSessionObject",
                "params": [{"qInfo": {"qType": "RIRA-Generator"},
                            "qHyperCubeDef": {
                                "qDimensions": dims, "qMeasures": [],
                                "qInitialDataFetch": [{"qTop": 0, "qLeft": 0, "qHeight": 1, "qWidth": n_cols}]
                            }}]
            }))

            obj_handle = None
            for _ in range(20):
                raw = await asyncio.wait_for(ws.recv(), timeout=15)
                msg = json.loads(raw)
                if msg.get("id") == 3:
                    if "error" in msg:
                        return None, f"CreateHyperCube error: {json.dumps(msg['error'])[:300]}"
                    if "result" in msg:
                        obj_handle = msg["result"]["qReturn"]["qHandle"]
                        break
                    else:
                        return None, f"CreateHyperCube odd response: {json.dumps(msg)[:500]}"

            if obj_handle is None:
                return None, "CreateHyperCube: no response received within timeout"

            progress_bar.progress(0.1, text="Data request created. Getting row count...")

            # 5. Get layout
            await ws.send(json.dumps({
                "jsonrpc": "2.0", "id": 4, "handle": obj_handle,
                "method": "GetLayout", "params": []
            }))

            total_rows = None
            for _ in range(20):
                raw = await asyncio.wait_for(ws.recv(), timeout=15)
                msg = json.loads(raw)
                if msg.get("id") == 4:
                    if "error" in msg:
                        return None, f"GetLayout error: {json.dumps(msg['error'])[:300]}"
                    if "result" in msg:
                        total_rows = msg["result"]["qLayout"]["qHyperCube"]["qSize"]["qcy"]
                        break

            if total_rows is None:
                return None, "GetLayout: no response received"
            if total_rows == 0:
                return None, "No data returned from Qlik (check if 'Approved' records exist for Non-Seed Care)"

            progress_bar.progress(0.15, text=f"Found {total_rows:,} rows (Approved only). Fetching data...")

            # 6. Paginate
            PAGE_SIZE = 500
            all_rows = []
            pages = (total_rows + PAGE_SIZE - 1) // PAGE_SIZE
            msg_id = 10

            for page in range(pages):
                top = page * PAGE_SIZE
                await ws.send(json.dumps({
                    "jsonrpc": "2.0", "id": msg_id, "handle": obj_handle,
                    "method": "GetHyperCubeData",
                    "params": ["/qHyperCubeDef", [{"qTop": top, "qLeft": 0,
                                "qHeight": min(PAGE_SIZE, total_rows - top), "qWidth": n_cols}]]
                }))

                for _ in range(20):
                    raw = await asyncio.wait_for(ws.recv(), timeout=30)
                    msg = json.loads(raw)
                    if msg.get("id") == msg_id:
                        if "error" in msg:
                            return None, f"GetData page {page} error: {json.dumps(msg['error'])[:200]}"
                        matrix = msg["result"]["qDataPages"][0]["qMatrix"]
                        all_rows.extend(matrix)
                        break

                msg_id += 1
                progress_bar.progress(
                    0.15 + 0.85 * min((page + 1) / pages, 1.0),
                    text=f"Fetching... {len(all_rows):,} / {total_rows:,} rows"
                )

            # 7. Convert to DataFrame
            data = [{fields[i]: cell.get("qText", "") for i, cell in enumerate(row)} for row in all_rows]
            df = pd.DataFrame(data)
            df.attrs["last_reload_time"] = last_reload_time
            return df, None


    except asyncio.TimeoutError:
        return None, "Connection timeout. Qlik did not respond in time."
    except KeyError as e:
        return None, f"KeyError: {e}. Qlik response structure unexpected."
    except Exception as e:
        return None, f"{type(e).__name__}: {str(e)[:400]}"

def fetch_qlik_sync(token, progress_bar):
    """Synchronous wrapper for async Qlik fetch."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    return loop.run_until_complete(fetch_qlik_data(token, progress_bar))


# ==============================================================================
#  FIXED BUSINESS RULES - UNIT CONVERSION TABLE (Reliance #43988)
# ==============================================================================

CONVERSION_FACTORS = {
    "kg/ha": 1000,
    "l/ha": 1000,
    "ml/ha": 1,
    "ml/m\u00b2": 10000,
    "g/m\u00b2": 10000,
    "oz (us)/acre": 73.07,
    "oz (imperial)/acre": 70.05,
    "lb/acre": 1120.85,
    "ml/100 m\u00b2": 100,
    "pints (us)/acre": 1168.77,
    "pints (imperial)/acre": 1169.33,
    "g/100 m\u00b2": 100,
    "kg/m\u00b2": 10000000,
    "oz (us)/1000 ft\u00b2": 3142.18,
    "oz (imperial)/1000 ft\u00b2": 3051.21,
    "gal (us)/acre": 9352.61,
    "gal (imperial)/acre": 9353.54,
    "gal/1000 row feet": 163164.22,
    "oz/short ton commodity": 12.87,
    "oz (us)/ft\u00b2": 318208.1,
    "oz (imperial)/ft\u00b2": 3051974,
    "ml/metric ton (1000kg) commodity": 1,
    "g/metric ton (1000kg) commodity": 1,
    "ml/1000 m3": 0.1,
    "kg/1000 kg commodity": 1000,
    "l/1000 kg commodity": 1000,
    "l/m\u00b2": 10000000,
    "ml/cm dbh": 3000,
    "g/1000 m3": 1000,
    "ml/m3": 10000,
    "kg/acre": 2471.05,
    "ml/1000 m\u00b2": 10,
    "ml/mu": 15,
    "g/mu": 15,
    "g/ha": 1,
}


def convert_to_gha(value, uom):
    """Convert a rate value to g/ha using the conversion table (Reliance #43988)."""
    if pd.isna(value) or pd.isna(uom):
        return np.nan
    try:
        value = float(value)
    except (ValueError, TypeError):
        return np.nan
    uom_lower = str(uom).strip().lower()
    factor = CONVERSION_FACTORS.get(uom_lower)
    if factor is None:
        return np.nan
    return value * factor

# ==============================================================================
#  CROP EQUIVALENTS MAPPING (Reliance #43922)
# ==============================================================================

CROP_EQUIVALENTS = {
    "ACACIA": "Fruit Tree",
    "AÃ‡AÃ": "Berry Crops",
    "ACEROLA": "Pome Fruit",
    "ACORN SQUASH": "Pumpkin/Squash",
    "ACTINIDIA ARGUTA": "Kiwi",
    "ADZUKI BEAN": "Beans (Species)",
    "ADZUKI BEAN (DRY)": "Beans (Species)",
    "AFRICAN EGGPLANT": "Aubergine/Eggplant",
    "AFRICAN MARIGOLD": "Ornamentals",
    "AFRICAN NUT TREE": "Fruit Tree",
    "AGRIMONY": "Ornamentals",
    "AJI": "Bell Pepper",
    "ALDER": "Shrub",
    "ALFAFA": "Rape/ Oilseed Rape",
    "ALFAFA FORAGE (GREEN)": "Rape/ Oilseed Rape",
    "ALFAFA, SEED": "Uncultivated Soil",
    "ALLIGATOR COCOA": "Cocoa/Cacao",
    "ALLIUM": "Garlic",
    "ALMOND": "Almond",
    "ALMONDS": "Almond",
    "ALSTROEMERIA": "Ornamentals",
    "AMARANTH": "Vegetable",
    "AMARANTH (LEAFY AMARANTH, CHINESE SPINACH, TAMPALA)": "Spinach",
    "AMARANTH, CHINESE": "Spinach",
    "AMARELLO CHERRY": "Stone Fruit",
    "AMENITY GRASSLAND": "Turf",
    "AMERICAN AGAVE": "Ornamentals",
    "AMERICAN BITTERSWEET": "Ornamentals",
    "AMERICAN GINSENG": "Ginseng",
    "AMERICAN GRAPE": "Grape",
    "AMERICAN OILPALM": "Oil Palm",
    "AMERICAN RED RASPBERRY": "Berry Crops",
    "AMERICAN STRAWBERRY": "Strawberries",
    "AMERICAN UPLAND COTTON/JUTE": "Cotton/Jute",
    "AMERICAN WHITE ELM": "Forestry",
    "AMUR RIVER GRAPE": "Grape",
    "ANACARDIACEAE": "Ornamentals",
    "ANDEAN BLACKBERRY": "Berry Crops",
    "ANDES BLACK RASPBERRY": "Berry Crops",
    "ANGELICA": "Ornamentals",
    "ANGELICA, INCLUDING GARDEN ANGELICA": "Ornamentals",
    "ANIMAL FEED, NONGRASS, GROUP 18": "Uncultivated Soil",
    "ANIMALS (GENERAL)": "Uncultivated Soil",
    "ANISE": "Vegetable",
    "ANNONACEAE": "Ornamentals",
    "ANNUAL RYEGRASS": "Perennial Ryegrass",
    "ANNUAL SAVORY": "Vegetable",
    "ANTHURIUM": "Ornamentals",
    "APPLE": "Apple",
    "APRICOT": "Apricot",
    "ARABIAN COFFEE": "Coffee",
    "ARABLE CROPS": "Uncultivated Soil",
    "ARABLE LAND": "Uncultivated Soil",
    "ARALIA": "Ornamentals",
    "ARAUCARIA ANGUSTIFOLIA": "Forestry",
    "ARCTIC BLACKBERRY": "Berry Crops",
    "ARECACEAE": "Ornamentals",
    "ARGYRANTHEMUM FRUTESCENS": "Ornamentals",
    "AROMATIC HERBS": "Vegetable",
    "ARONIA BERRY": "Berry Crops",
    "ARRACACHA": "Vegetable",
    "ARRACACIA XANTHORRHIZA": "Vegetable",
    "ARROWHEAD": "Vegetable",
    "ARROWROOT": "Vegetable",
    "ARTICHOKE": "Artichoke",
    "ARTICHOKE, CHINESE": "Artichoke",
    "ARTICHOKE, GLOBE": "Artichoke",
    "ARTICHOKE, JERUSALEM": "Artichoke",
    "ARUGULA": "Vegetable",
    "ARUGULA (ROQUETTE)": "Vegetable",
    "ASATSUKI": "Onion",
    "ASH": "Olive",
    "ASIATIC COTTON/JUTE": "Cotton/Jute",
    "ASPARAGUS": "Asparagus",
    "ASPARAGUS BEAN": "Beans (Species)",
    "ASPARAGUS BROCCOLI": "Asparagus",
    "ASTER": "Ornamentals",
    "ASTER, INDIAN": "Ornamentals",
    "ASTERACEAE": "Ornamentals",
    "AUBERGINE": "Aubergine/Eggplant",
    "AVOCADO": "Avocado",
    "AZALEA": "Ornamentals",
    "AZAROLE": "Pome Fruit",
    "BAHIAGRASS": "Perennial Ryegrass",
    "BALLHEAD ONION": "Onion",
    "BALLOONFLOWER": "Ornamentals",
    "BALM LEAVES": "Vegetable",
    "BALSAM APPLE": "Curcubit",
    "BALSAM PEAR": "Curcubit",
    "BAMBOO": "Vegetable",
    "BANANA": "Banana",
    "BANANA SHRUB": "Shrub",
    "BANITI": "Fruit Tree",
    "BARBADINE": "Fruit Tree",
    "BARBADOS CHERRY": "Cherry",
    "BARLEY": "Barley",
    "BASIL": "Vegetable",
    "BASILICUM": "Vegetable",
    "BATIKI BLUEGRASS": "Perennial Ryegrass",
    "BEAN": "Beans (Species)",
    "BEAN, BLACK TURTLE": "Beans (Species)",
    "BEAN, DRY": "Beans (Species)",
    "BEAN, DRY COMMON": "Beans (Species)",
    "BEAN, EDIBLE PODDED, SUCCULENT": "Beans (Species)",
    "BEAN, GARDEN": "Beans (Species)",
    "BEAN, GREEN": "Beans (Species)",
    "BEAN, VELVET": "Beans (Species)",
    "BEANS (DRY)": "Beans (Species)",
    "BEARBERRY": "Berry Crops",
    "BEE BALM": "Uncultivated Soil",
    "BEECHNUT": "Uncultivated Soil",
    "BEET": "Sugarbeet",
    "BEET LEAVES": "Sugarbeet",
    "BEET, FODDER": "Sugarbeet",
    "BEET, GARDEN": "Sugarbeet",
    "BEET, SUGAR": "Sugarbeet",
    "BEETROOT": "Sugarbeet",
    "BEETS": "Sugarbeet",
    "BEGONIA": "Ornamentals",
    "BELL PEPPER": "Bell Pepper",
    "BELLYACHE BUSH": "Uncultivated Soil",
    "BENTGRASS": "Turf",
    "BERMUDA GRASS": "Turf",
    "BERMUDAGRASS": "Turf",
    "BERRIES AND OTHER SMALL FRUITS": "Berry Crops",
    "BERRY AND SMALL FRUIT, SUBGROUP 13-7": "Berry Crops",
    "BERRY AND SMALL FRUIT, SUBGROUP 13-07A": "Berry Crops",
    "BERRY AND SMALL FRUIT, SUBGROUP 13-07B": "Berry Crops",
    "BERRY PLANTS": "Berry Crops",
    "BERSEEM CLOVER": "Clover",
    "BETA BEET": "Sugarbeet",
    "BETEL LEAF": "Uncultivated Soil",
    "BETEL LEAVES": "Uncultivated Soil",
    "BETEL VINE": "Uncultivated Soil",
    "BETELNUT": "Uncultivated Soil",
    "BETTE": "Vegetable",
    "BIG MARIGOLD": "Ornamentals",
    "BILBERRY": "Berry Crops",
    "BILBERRY, BOG": "Berry Crops",
    "BINGLEBERRY": "Berry Crops",
    "BIRD'S FOOT": "Uncultivated Soil",
    "BIRDSFOOT-TREFOIL": "Uncultivated Soil",
    "BITTER BALSAM APPLE": "Curcubit",
    "BITTER GOURD": "Vegetable",
    "BITTER MELON": "Vegetable",
    "BITTER ORANGE": "Citrus",
    "BLACK CHOKEBERRY": "Berry Crops",
    "BLACK CURRANT": "Berry Crops",
    "BLACK ELDER": "Berry Crops",
    "BLACK GRAM (GREEN PODS)": "Beans (Species)",
    "BLACK MULBERRY": "Berry Crops",
    "BLACK MUSTARD": "Brassicas",
    "BLACK RASPBERRY": "Berry Crops",
    "BLACK SALSIFY": "Vegetable",
    "BLACK SATIN BERRY": "Berry Crops",
    "BLACKBERRIES": "Berry Crops",
    "BLACKBERRY": "Berry Crops",
    "BLACKEYED PEA": "Beans (Species)",
    "BLACKJACK": "Uncultivated Soil",
    "BLATTA SP.": "Public Health",
    "BLATTELLA GERMANICA": "Public Health",
    "BLEACHED CELERY": "Celery",
    "BLUE AFRICAN VIOLET": "Ornamentals",
    "BLUE LUPIN": "Lupin",
    "BLUEBERRIES": "Berry Crops",
    "BLUEBERRY": "Berry Crops",
    "BLUEBERRY, HIGHBUSH": "Berry Crops",
    "BLUEBERRY, LOWBUSH": "Berry Crops",
    "BLUE-CROWN PASSION FLOWER": "Fruit",
    "BLUEGRASS": "Turf",
    "BORAGE": "Uncultivated Soil",
    "BORDER AND EMBANKMENT OF WAYS OR DITCHES": "Uncultivated Soil",
    "BOTTLE GOURD": "Cucurbit",
    "BOYSENBERRY": "Berry Crops",
    "BRACKEN": "Uncultivated Soil",
    "BRASSICA": "Brassicas",
    "BRASSICA (COLE OR CABBAGE) VEGETABLE, HEAD CABBAGE, FLOWERHEAD BRASSICAS": "Brassicas",
    "BRASSICA CAMPESTRIS": "Brassicas",
    "BRASSICA HEAD AND STEM VEGETABLE GROUP": "Brassicas",
    "BRASSICA JUNCEA": "Brassicas",
    "BRASSICA NAPUS": "Rape/ Oilseed Rape",
    "BRASSICA OLERACEA": "Brassicas",
    "BRASSICA RAPA SUBSP. OLEIFERA": "Rape/ Oilseed Rape",
    "BRASSICA SP.": "Brassicas",
    "BRASSICA, HEAD AND STEM, SUBGROUP 5A": "Brassicas",
    "BRASSICA, LEAFY GREENS, SUBGROUP 5B": "Brassicas",
    "BRASSICACEAE": "Brassicas",
    "BRAZIL NUT": "Hazelnut",
    "BRAZILIAN LUCERNE": "Rape/ Oilseed Rape",
    "BREAD WHEAT": "Wheat",
    "BREADFRUIT": "Fruit",
    "BREEDING AND SEED PRODUCTION": "Uncultivated Soil",
    "BRIAR ROSE": "Ornamentals",
    "BROAD BEAN": "Beans (Species)",
    "BROAD BEAN (FAVA BEAN)": "Beans (Species)",
    "BROADBEAKED MUSTARD": "Brassicas",
    "BROADLEAF FOREST TREES": "Forestry",
    "BROAD-LEAVED ENDIVE": "Chicory/Endive",
    "BROAD-LEAVED PLANTS": "Uncultivated Soil",
    "BROCCOLI": "Broccoli",
    "BROCCOLI RAAB": "Broccoli",
    "BROCCOLI, CHINESE": "Broccoli",
    "BROCCOLI, CHINESE (GAI LON)": "Broccoli",
    "BROME GRASS": "Turf",
    "BROMEGRASS": "Turf",
    "BROMPTON STOCK": "Ornamentals",
    "BROOM MILLET": "Cereal",
    "BROOMCORN MILLET": "Cereal",
    "BROWN SARSON": "Rape/ Oilseed Rape",
    "BRUSSELS SPROUTS": "Brassicas",
    "BUCKWHEAT": "Cereal",
    "BUFFALO CURRANT": "Berry Crops",
    "BULB IRIS": "Ornamentals",
    "BULB VEGETABLE": "Vegetable",
    "BULB VEGETABLE, EXCEPT FENNEL, BULB": "Vegetable",
    "BULBOUS CRANESBILL": "Ornamentals",
    "BULBOUS PLANTS": "Ornamentals",
    "BUPLEURUM CHINENSE": "Uncultivated Soil",
    "BURDOCK, GREATER OR EDIBLE": "Vegetable",
    "BURNET": "Uncultivated Soil",
    "BURNET-SAXIFRAGE": "Uncultivated Soil",
    "BUSH PUMPKIN": "Pumpkin/Squash",
    "BUSH SQUASH": "Pumpkin/Squash",
    "BUSHBERRY, SUBGROUP 13B": "Berry Crops",
    "BUSHWEED": "Uncultivated Soil",
    "BUTTER BEAN": "Beans (Species)",
    "BUTTERBUR": "Uncultivated Soil",
    "BUTTERNUT": "Hazelnut",
    "BUTTERNUT SQUASH": "Pumpkin/Squash",
    "CABBAGE": "Cabbage",
    "CABBAGE, CHINESE (NAPA)": "Cabbage",
    "CABBAGE, CHINESE MUSTARD (GAI CHOY)": "Cabbage",
    "CABBAGE, RED": "Cabbage",
    "CABBAGE, SAVOY": "Cabbage",
    "CABBAGES, HEAD": "Cabbage",
    "CACAO": "Cocoa/Cacao",
    "CACAO BEAN": "Cocoa/Cacao",
    "CACAO BEANS": "Cocoa/Cacao",
    "CALABAZA": "Pumpkin/Squash",
    "CALABRESE": "Broccoli",
    "CALAMONDIN": "Citrus",
    "CALENDULA": "Ornamentals",
    "CALIFORNIA BLACKBERRY": "Berry Crops",
    "CAMELINA": "Rape/ Oilseed Rape",
    "CAMELINA SATIVA": "Rape/ Oilseed Rape",
    "CAMOMILE OR CHAMOMILE": "Uncultivated Soil",
    "CANARYGRASS": "Cereal",
    "CANEBERRY": "Berry Crops",
    "CANEBERRY, SUBGROUP 13A": "Berry Crops",
    "CANISTEL": "Fruit",
    "CANNA, EDIBLE": "Vegetable",
    "CANOLA": "Rape/ Oilseed Rape",
    "CANTALOUPE": "Melon",
    "CAPE GOOSEBERRY": "Berry Crops",
    "CAPSICUM": "Bell Pepper",
    "CAQUI": "Fruit",
    "CARAMBOLA": "Fruit",
    "CARAWAY": "Uncultivated Soil",
    "CARAWAY SEED": "Uncultivated Soil",
    "CARDAMOM": "Uncultivated Soil",
    "CARDAMOM SEED": "Uncultivated Soil",
    "CARDAMON": "Uncultivated Soil",
    "CARDOON": "Vegetable",
    "CARNATION": "Ornamentals",
    "CAROB BEAN": "Uncultivated Soil",
    "CAROSELLA": "Vegetable",
    "CARROT": "Carrot",
    "CASABA": "Melon",
    "CASHEW": "Uncultivated Soil",
    "CASHEW APPLE": "Fruit",
    "CASSAVA": "Cassava / Manioc",
    "CASSAVA, BITTER AND SWEET": "Cassava / Manioc",
    "CASTOR OIL PLANT": "Uncultivated Soil",
    "CASTORBEAN": "Uncultivated Soil",
    "CATJANG": "Beans (Species)",
    "CAT'S WHISKERS": "Uncultivated Soil",
    "CAT-TAIL MILLET": "Cereal",
    "CAULIFLOWER": "Cauliflower",
    "CAVALO BROCCOLO": "Broccoli",
    "CEDAR": "Forestry",
    "CELERIAC": "Celeriac",
    "CELERIAC (CELERY ROOT)": "Celeriac",
    "CELERY": "Celery",
    "CELERY LETTUCE": "Lettuce",
    "CELERY, CHINESE": "Celery",
    "CELERY, DRIED LEAVES": "Celery",
    "CELERY, ROOTED": "Celeriac",
    "CELERY, SEED": "Celery",
    "CELOSIA": "Ornamentals",
    "CELTUCE": "Lettuce",
    "CEMENT": "Uncultivated Soil",
    "CENTIPEDEGRASS": "Turf",
    "CEREAL, FLOUR": "Cereal",
    "CEREALS": "Cereal",
    "CEREALS (STORED)": "Cereal",
    "CEREALS STUBBLE FIELD": "Cereal",
    "CHAIN-LINK CACTUS": "Ornamentals",
    "CHAM-CHWI": "Vegetable",
    "CHAM-NAMUL": "Vegetable",
    "CHAMOMILLE": "Uncultivated Soil",
    "CHARD BEET": "Sugarbeet",
    "CHAYA": "Vegetable",
    "CHAYOTE": "Vegetable",
    "CHAYOTE (FRUIT)": "Vegetable",
    "CHAYOTE ROOT": "Vegetable",
    "CHEMICAL FALLOW": "Uncultivated Soil",
    "CHERIMOYA": "Fruit",
    "CHEROKEE BLACKBERRY": "Berry Crops",
    "CHERRIES": "Cherry",
    "CHERRY": "Cherry",
    "CHERRY LAUREL": "Cherry",
    "CHERRY TOMATO": "Tomato",
    "CHERRY, BLACK": "Cherry",
    "CHERRY, SOUR": "Cherry",
    "CHERRY, SWEET": "Cherry",
    "CHERRY, SWEET OR CHERRY, TART, SUBGROUP 12-12 A": "Cherry",
    "CHERRY, TART": "Cherry",
    "CHERVIL": "Vegetable",
    "CHERVIL, FRESH LEAVES": "Vegetable",
    "CHESS BROMEGRASS": "Turf",
    "CHESTERBERRY": "Berry Crops",
    "CHESTNUT": "Hazelnut",
    "CHESTNUTS": "Hazelnut",
    "CHEYENNE BLACKBERRY": "Berry Crops",
    "CHICKPEA": "Beans (Species)",
    "CHICKPEA (GARBANZO BEAN)": "Beans (Species)",
    "CHICK-PEA (GREEN PODS)": "Beans (Species)",
    "CHICORY": "Chicory/Endive",
    "CHICORY, ROOTS": "Chicory/Endive",
    "CHILEAN GUAVA": "Fruit",
    "CHILEAN STRAWBERRY": "Strawberries",
    "CHILE": "Bell Pepper",
    "CHILI": "Bell Pepper",
    "CHILI PEPPER": "Bell Pepper",
    "CHILLI": "Bell Pepper",
    "CHINA ROSE": "Ornamentals",
    "CHINESE CABBAGE": "Cabbage",
    "CHINESE CABBAGE (TYPE PE-TSAI)": "Cabbage",
    "CHINESE CHARD": "Vegetable",
    "CHINESE FAIRY GRASS": "Uncultivated Soil",
    "CHINESE FLOWERING CABBAGE": "Cabbage",
    "CHINESE KALE": "Brassicas",
    "CHINESE MUSTARD": "Brassicas",
    "CHINESE OKRA": "Curcubit",
    "CHINESE PEAR": "Pear",
    "CHINESE PEONY": "Ornamentals",
    "CHINESE SPINACH": "Spinach",
    "CHINESE SQUASH": "Pumpkin/Squash",
    "CHINESE STRAWBERRY TREE": "Fruit Tree",
    "CHINESE SWEET CANE": "Sugarcane",
    "CHINESE TURNIP": "Vegetable",
    "CHINESE WAXGOURD": "Cucurbit",
    "CHINOLI": "Fruit",
    "CHIPILIN": "Vegetable",
    "CHIVE": "Vegetable",
    "CHIVE, CHINESE": "Vegetable",
    "CHIVE, FRESH LEAVES": "Vegetable",
    "CHIVE, LEAVES": "Vegetable",
    "CHIVES": "Vegetable",
    "CHOKEBERRY": "Berry Crops",
    "CHOP SUEY GREENS": "Vegetable",
    "CHOY SUM": "Vegetable",
    "CHRISTMAS TREES": "Forestry",
    "CHRYSANTHEMUM": "Ornamentals",
    "CHRYSANTHEMUM SP.": "Ornamentals",
    "CHRYSANTHEMUM, EDIBLE LEAVED": "Vegetable",
    "CHRYSANTHEMUM, GARLAND": "Ornamentals",
    "CHUFA": "Vegetable",
    "CICER MILK VETCH": "Uncultivated Soil",
    "CICHORIUM INTYBUS": "Chicory/Endive",
    "CILANTRO": "Vegetable",
    "CILANTRO, FRESH LEAVES": "Vegetable",
    "CINQUEFOIL": "Uncultivated Soil",
    "CISTUS SP.": "Ornamentals",
    "CITRON": "Citrus",
    "CITRON MELON": "Melon",
    "CITRON, CITRUS": "Citrus",
    "CITRUL": "Watermelon",
    "CITRUS": "Citrus",
    "CITRUS FRUIT CROPS": "Citrus",
    "CITRUS FRUITS": "Citrus",
    "CITRUS HYBRIDS": "Citrus",
    "CITRUS SP.": "Citrus",
    "CLARY SAGE": "Uncultivated Soil",
    "CLEMENTINE": "Citrus",
    "CLIMBING FRENCH BEAN": "Beans (Species)",
    "CLIMBING ROSES": "Ornamentals",
    "CLOUDBERRY": "Berry Crops",
    "CLOVE": "Uncultivated Soil",
    "CLOVER": "Clover",
    "CLOVER PLANTS": "Clover",
    "CLOVER, FORAGE": "Clover",
    "CLOVER, RED": "Clover",
    "CLOVER, SUB": "Clover",
    "CNIDIUM": "Uncultivated Soil",
    "COASTAL SWEET PEPPER-BUSH": "Shrub",
    "COCA-BUSH": "Uncultivated Soil",
    "COCKROACHES": "Public Health",
    "COCKSFOOT": "Turf",
    "COCKSPUR CORAL TREE": "Uncultivated Soil",
    "COCONA": "Fruit",
    "COCONUT": "Fruit",
    "COCOYAM": "Vegetable",
    "CODONOPSIS": "Uncultivated Soil",
    "COFFEA SP.": "Coffee",
    "COFFEE": "Coffee",
    "COFFEE BEANS": "Coffee",
    "COFFEE CHICORY": "Chicory/Endive",
    "COLE CROP": "Brassicas",
    "COLLARDS": "Brassicas",
    "COLOCYNTH": "Cucurbit",
    "COLZA": "Rape/ Oilseed Rape",
    "COMMON ANGELICA": "Uncultivated Soil",
    "COMMON ASH": "Forestry",
    "COMMON BANANA": "Banana",
    "COMMON BENTGRASS": "Turf",
    "COMMON BLACKBERRY": "Berry Crops",
    "COMMON BORAGE": "Uncultivated Soil",
    "COMMON BOX": "Shrub",
    "COMMON CACAO": "Cocoa/Cacao",
    "COMMON CARAWAY": "Uncultivated Soil",
    "COMMON COCONUT PALM": "Fruit Tree",
    "COMMON CORIANDER": "Vegetable",
    "COMMON CORN SALAD": "Corn",
    "COMMON COWPEA": "Beans (Species)",
    "COMMON DANDELION": "Ornamentals",
    "COMMON DATE PALM": "Fruit Tree",
    "COMMON ELDER": "Shrub",
    "COMMON ELM": "Forestry",
    "COMMON EVENING PRIMROSE": "Uncultivated Soil",
    "COMMON FENNEL": "Fennel",
    "COMMON FENUGREEK": "Uncultivated Soil",
    "COMMON FIG": "Fruit",
    "COMMON FLAX": "Linseed/Flax",
    "COMMON GINGER": "Ginseng",
    "COMMON HEMP": "Uncultivated Soil",
    "COMMON HOP": "Hops",
    "COMMON JUJUBE": "Fruit",
    "COMMON LEEK": "Leek",
    "COMMON MEDLAR": "Fruit",
    "COMMON NASTURTIUM": "Ornamentals",
    "COMMON OAT": "Oat",
    "COMMON PASSIONFRUIT": "Fruit",
    "COMMON PEAR": "Pear",
    "COMMON PERILLA": "Uncultivated Soil",
    "COMMON PINK": "Ornamentals",
    "COMMON POPPY": "Uncultivated Soil",
    "COMMON PURSLANE": "Vegetable",
    "COMMON REED": "Uncultivated Soil",
    "COMMON RICE": "Rice",
    "COMMON SAGE": "Shrub",
    "COMMON SALSIFY": "Vegetable",
    "COMMON SORREL": "Vegetable",
    "COMMON ST. JOHNSWORT": "Uncultivated Soil",
    "COMMON SUGARCANE": "Sugarcane",
    "COMMON SUNFLOWER": "Sunflower",
    "COMMON SYCAMORE": "Forestry",
    "COMMON THYME": "Uncultivated Soil",
    "COMMON VALERIAN": "Uncultivated Soil",
    "COMMON VETCH": "Uncultivated Soil",
    "COMMON WALNUT": "Hazelnut",
    "COMMON WINTERCRESS": "Vegetable",
    "COMMON WORMWOOD": "Uncultivated Soil",
    "CONIFEROUS FOREST TREES": "Forestry",
    "CONIFEROUS PLANTS": "Forestry",
    "CONIFEROUS TREES": "Forestry",
    "COOKING PEPPER": "Bell Pepper",
    "COOPER GLYCINE": "Uncultivated Soil",
    "CORDYLINE FRUTICOSA": "Ornamentals",
    "CORIANDER": "Vegetable",
    "CORN": "Corn",
    "CORN FLAG": "Ornamentals",
    "CORN SALAD": "Corn",
    "CORN, FIELD": "Corn",
    "CORN, POP": "Corn",
    "CORN, SEED": "Corn",
    "CORN, SWEET": "Corn",
    "CORYBERRY": "Berry Crops",
    "CORYLUS": "Hazelnut",
    "COS LETTUCE": "Lettuce",
    "COSMOS": "Ornamentals",
    "COTTON/JUTE": "Cotton/Jute",
    "COTTON/JUTE PLANT": "Cotton/Jute",
    "COTTON/JUTESEED, SUBGROUP 20C": "Cotton/Jute",
    "COW CABBAGE": "Brassicas",
    "COWPEA": "Beans (Species)",
    "COWPEA (DRY)": "Beans (Species)",
    "CRAB APPLE": "Apple",
    "CRABAPPLE": "Apple",
    "CRAB-APPLE": "Apple",
    "CRAMBE": "Vegetable",
    "CRANBERRY": "Berry Crops",
    "CRANESBILL": "Ornamentals",
    "CREAMY BUTTERBUR": "Vegetable",
    "CRENSHAW MELON": "Melon",
    "CRESS, GARDEN": "Vegetable",
    "CRESS, UPLAND": "Vegetable",
    "CRIMSON CLOVER": "Clover",
    "CROP GROUP: FRUITING VEG (CG8 & 8-9) (PMRA)": "Vegetable",
    "CROP GROUP: STONE FRUIT (CG12 & 12-09) (PMRA)": "Stone Fruit",
    "CROP GROUP: POME FRUIT (CG11 & 11-09) (PMRA)": "Pome Fruit",
    "CROP PLANTS": "Vegetable",
    "CROWDER PEA": "Pea",
    "CUCUMBER": "Cucumber",
    "CUCUMBER TREE": "Fruit Tree",
    "CUCUMBER, CHINESE": "Cucumber",
    "CUCURBITA": "Cucurbit",
    "CUCURBITA MAXIMA": "Cucurbit",
    "CUCURBITA MELOPEPO": "Cucurbit",
    "CUCURBITA PEPO": "Cucurbit",
    "CUCURBITA PEPO VAR. OLEIFERA": "Cucurbit",
    "CUCURBITACEAE": "Cucurbit",
    "CULICOIDES SP.": "Public Health",
    "CUMIN": "Vegetable",
    "CUMIN SEED": "Vegetable",
    "CUPHEA": "Ornamentals",
    "CUPUACÃš": "Fruit",
    "CURLED KITCHEN KALE": "Brassicas",
    "CURLED MALLOW": "Vegetable",
    "CURLY KALE": "Brassicas",
    "CURRANT": "Berry Crops",
    "CURRANT BUSH": "Berry Crops",
    "CURRANT TOMATO": "Tomato",
    "CURRANT, BLACK": "Berry Crops",
    "CURRANT, RED": "Berry Crops",
    "CURRANT, RED, WHITE": "Berry Crops",
    "CURRANTS, BLACK, RED, WHITE": "Berry Crops",
    "CUSTARD APPLE": "Fruit",
    "CUT-FLOWER PLANTS": "Ornamentals",
    "CUTTING LETTUCE": "Lettuce",
    "CYCLAMEN": "Ornamentals",
    "CYPRESS": "Forestry",
    "DAFFODIL": "Ornamentals",
    "DAHLIA": "Ornamentals",
    "DAISY": "Ornamentals",
    "DANDELION": "Ornamentals",
    "DARROWBERRY": "Berry Crops",
    "DASHEEN": "Vegetable",
    "DATE": "Fruit",
    "DAYLILY": "Ornamentals",
    "DECIDUOUS FRUIT": "Deciduous Fruit",
    "DECIDUOUS WOODY PLANTS": "Deciduous Fruit",
    "DENT CORN": "Corn",
    "DEVIL'S PINCUSHION": "Uncultivated Soil",
    "DEWBERRIES (INCLUDING BOYSENBERRY AND LOGANBERRY)": "Berry Crops",
    "DEWBERRY": "Berry Crops",
    "DIANTHUS": "Ornamentals",
    "DILL": "Vegetable",
    "DIOSCOREA TRIFIDA": "Vegetable",
    "DIPLOTAXIS TENUIFOLIA": "Vegetable",
    "DIRECT SEEDED ONION": "Onion",
    "DIRKSEN THORNLESS BERRY": "Berry Crops",
    "DITCH": "Uncultivated Soil",
    "DOCK": "Vegetable",
    "DOGWOOD": "Ornamentals",
    "DOL-NAMMUL": "Vegetable",
    "DRAGON FRUIT": "Fruit",
    "DRIED SHELLED AND SUCCULENT BEANS, EXCEPT COWPEA": "Beans (Species)",
    "DRY BEAN": "Beans (Species)",
    "DRY-SEEDED PADDY RICE": "Rice",
    "DUBOISIA": "Uncultivated Soil",
    "DURIAN": "Fruit",
    "DURUM WHEAT": "Wheat",
    "DURUM WHEAT (SPRING)": "Wheat",
    "DURUM WHEAT (WINTER)": "Wheat",
    "DUTCH CLOVER": "Clover",
    "DUTCH ELM": "Forestry",
    "DWARF BANANA": "Banana",
    "DWARF FRENCH BEAN": "Beans (Species)",
    "EARLY WINTERCRESS": "Vegetable",
    "EARTH PEA": "Pea",
    "EAST INDIAN MATTINGGRASS": "Turf",
    "EASTER LILY": "Ornamentals",
    "EBOLO": "Vegetable",
    "ECHIUM": "Uncultivated Soil",
    "EDIBLE BURDOCK": "Vegetable",
    "EDIBLE GOURD": "Cucurbit",
    "EDIBLE PODDED PEA": "Pea",
    "EGG PLANT": "Aubergine/Eggplant",
    "EGGPLANT": "Aubergine/Eggplant",
    "EINKORN WHEAT": "Wheat",
    "ELDERBERRY": "Berry Crops",
    "EMMER": "Wheat",
    "ENDIVE": "Chicory/Endive",
    "ENDIVE (ESCAROLE)": "Chicory/Endive",
    "ENDIVE, BROAD OR PLAIN LEAVED": "Chicory/Endive",
    "ENGLISH OAK": "Forestry",
    "ESCAROLE": "Chicory/Endive",
    "ESTRAGON": "Vegetable",
    "ETL_BLANK": "Uncultivated Soil",
    "ETL_NOT_MAPPED": "Uncultivated Soil",
    "ETL_UNDETERMINED": "Uncultivated Soil",
    "EUCALYPTUS": "Forestry",
    "EUONYMUS EUROPAEUS": "Ornamentals",
    "EUPHORBIA": "Ornamentals",
    "EUROPEAN BARBERRY": "Berry Crops",
    "EUROPEAN BEECH": "Forestry",
    "EUROPEAN GRAPE": "Grape",
    "EUROPEAN PLUM": "Plum",
    "EUROPEAN SILVER FIR": "Forestry",
    "EUSTOMA": "Ornamentals",
    "EVENING PRIMROSE": "Uncultivated Soil",
    "EXACUM": "Ornamentals",
    "FABA BEAN": "Beans (Species)",
    "FABACEAE": "Beans (Species)",
    "FALLOW": "Uncultivated Soil",
    "FAMEFLOWER": "Ornamentals",
    "FARKLEBERRY": "Berry Crops",
    "FEATHER COCKSCOMB": "Ornamentals",
    "FEED PROCESSING AND STORAGE AREAS": "Uncultivated Soil",
    "FEIJOA": "Fruit",
    "FENNEL": "Fennel",
    "FENNEL, FLORENCE (FINOCHIO)": "Fennel",
    "FENNEL, FLORENCE, FRESH LEAVES AND STALK": "Fennel",
    "FENUGREEK": "Vegetable",
    "FESCUE": "Turf",
    "FEVERFEW": "Uncultivated Soil",
    "FIELD BEAN": "Beans (Species)",
    "FIELD CABBAGE": "Cabbage",
    "FIELD CROPS": "Vegetable",
    "FIELD PEA": "Pea",
    "FIG": "Fruit",
    "FILBERT": "Hazelnut",
    "FILBERT (HAZELNUT)": "Hazelnut",
    "FIR": "Forestry",
    "FLAX": "Linseed/Flax",
    "FLAX SEED": "Linseed/Flax",
    "FLORENCE FENNEL": "Fennel",
    "FLORICULTURE CROPS": "Ornamentals",
    "FLORISTS' CHRYSANTHEMUM": "Ornamentals",
    "FLOWER SEEDS": "Ornamentals",
    "FLOWER, EDIBLE": "Ornamentals",
    "FLOWERHEAD BRASSICA CROPS": "Brassicas",
    "FLOWERHEAD BRASSICAS (INCLUDES BROCCOLI: BROCCOLI, CHINESE AND CAULIFLOWER)": "Brassicas",
    "FODDER BEET": "Sugarbeet",
    "FODDER LEGUMES PLANTS": "Vegetable",
    "FORAGE TURNIP": "Vegetable",
    "FOREST (AFFORESTATION)": "Forestry",
    "FOREST (CLOSED WOOD)": "Forestry",
    "FOREST (NURSERY)": "Forestry",
    "FOREST TREES": "Forestry",
    "FOXBERRY": "Berry Crops",
    "FOXGLOVE": "Ornamentals",
    "FRENCH BEAN": "Beans (Species)",
    "FRUIT": "Fruit",
    "FRUIT PLANTS": "Fruit",
    "FRUIT TREE PLANTS": "Fruit Tree",
    "FRUIT TREES": "Fruit Tree",
    "FRUIT VEGETABLE PLANTS": "Vegetable",
    "FRUIT, BERRIES, GROUP 13": "Berry Crops",
    "FRUIT, CITRUS, EXCEPT MANDARIN": "Citrus",
    "FRUIT, CITRUS, GROUP 10": "Citrus",
    "FRUIT, CITRUS, GROUP 10-10": "Citrus",
    "FRUIT, POME": "Pome Fruit",
    "FRUIT, POME, GROUP 11": "Pome Fruit",
    "FRUIT, POME, GROUP 11-10": "Pome Fruit",
    "FRUIT, STONE, EXCEPT PLUM, PRUNE, DRIED": "Stone Fruit",
    "FRUIT, STONE, GROUP 12": "Stone Fruit",
    "FRUIT, STONE, GROUP 12-12": "Stone Fruit",
    "FRUITING VEGETABLE": "Vegetable",
    "FRUITING VEGETABLE OTHER THAN CUCURBITS": "Vegetable",
    "FRUITING VEGETABLE, CUCURBITS": "Cucurbit",
    "FRUITS": "Fruit",
    "FRUITS (EXCEPT AS OTHERWISE LISTED)": "Fruit",
    "FRUITS AND TREE NUTS": "Fruit",
    "FUCHSIA": "Ornamentals",
    "FUKI": "Vegetable",
    "GARDEN ANGELICA": "Vegetable",
    "GARDEN ASPARAGUS": "Asparagus",
    "GARDEN BEAN": "Beans (Species)",
    "GARDEN CHERVIL": "Vegetable",
    "GARDEN CRESS": "Vegetable",
    "GARDEN DAHLIA": "Ornamentals",
    "GARDEN HUCKLEBERRY": "Berry Crops",
    "GARDEN ONION": "Onion",
    "GARDEN PANSY": "Ornamentals",
    "GARDEN PARSLEY": "Parsley",
    "GARDEN PEA": "Pea",
    "GARDEN PETUNIA": "Ornamentals",
    "GARDEN PURSLANE": "Vegetable",
    "GARDEN RADISH": "Radish",
    "GARDEN ROSEMARY": "Vegetable",
    "GARDEN RUE": "Vegetable",
    "GARDEN SORREL": "Vegetable",
    "GARDEN STRAWBERRY": "Strawberries",
    "GARDEN THYME": "Vegetable",
    "GARDEN TURNIP": "Vegetable",
    "GARDEN VERBENA": "Ornamentals",
    "GARDEN WALL-ROCKET": "Vegetable",
    "GARLAND CHRYSANTHEMUM": "Vegetable",
    "GARLIC": "Garlic",
    "GARLIC, BULB": "Garlic",
    "GARLIC, GREAT HEADED": "Garlic",
    "GARLIC, SERPENT, BULB": "Garlic",
    "GERANIUM": "Ornamentals",
    "GERBERA": "Ornamentals",
    "GERMAN CELERY": "Celery",
    "GHERKIN": "Cucumber",
    "GHERKIN CROPS": "Cucumber",
    "GIANT GARLIC": "Garlic",
    "GIANT PUMPKIN": "Pumpkin/Squash",
    "GINGER": "Ginseng",
    "GINGER, ROOT": "Ginseng",
    "GINKGO": "Ornamentals",
    "GINSENG": "Ginseng",
    "GLADIOLUS": "Ornamentals",
    "GLOSSY ABELIA": "Ornamentals",
    "GLYCINE SOJA": "Soybean",
    "GOJI BERRY": "Berry Crops",
    "GOLDEN PERSHAW MELON": "Melon",
    "GOLD-OF-PLEASURE": "Rape/ Oilseed Rape",
    "GOLF FAIRWAY TURF": "Turf",
    "GOLF GREEN TURF": "Turf",
    "GOLF TEE TURF": "Turf",
    "GOOD KING HENRY": "Vegetable",
    "GOOSEBERRY": "Gooseberry",
    "GOURD": "Cucurbit",
    "GOURD, EDIBLE": "Cucurbit",
    "GRAIN LUPIN": "Lupin",
    "GRAIN, CEREAL, FORAGE, FODDER AND STRAW, GROUP 16": "Cereal",
    "GRAIN, CEREAL, GROUP 15": "Cereal",
    "GRAIN, FORAGE AND STOVER": "Cereal",
    "GRAPE": "Grape",
    "GRAPE (VITIS SPP.)": "Grape",
    "GRAPE, WINE": "Grape",
    "GRAPEFRUIT": "Citrus",
    "GRAPEFRUIT, SUBGROUP 10": "Citrus",
    "GRAPES": "Grape",
    "GRAPEVINE": "Grape",
    "GRASS": "Turf",
    "GRASS SEED": "Turf",
    "GRASS, FORAGE": "Turf",
    "GRASS, FORAGE, FODDER AND HAY, GROUP 17": "Turf",
    "GRASS, GRAMA": "Turf",
    "GRASS, HAY": "Turf",
    "GRASS, PASTURE": "Turf",
    "GRASS, RANGELAND": "Turf",
    "GRASS, SEED, STRAW": "Turf",
    "GRASS, ST. AUGUSTINE": "Turf",
    "GRASS, ZOYSIA": "Turf",
    "GRASSES": "Turf",
    "GRASSLAND": "Turf",
    "GRASSLAND NOT USED IN AGRICULTURE": "Uncultivated Soil",
    "GREAT SNAPDRAGON": "Ornamentals",
    "GRECIAN FOXGLOVE": "Ornamentals",
    "GREEN ONION": "Onion",
    "GREEN PEA": "Pea",
    "GREENGAGE": "Plum",
    "GREY PEA": "Pea",
    "GROUNDCHERRY": "Vegetable",
    "GROUNDNUT": "Peanut",
    "GUAR": "Vegetable",
    "GUARANA": "Fruit",
    "GUAVA": "Fruit",
    "GYNURA BICOLOR": "Vegetable",
    "GYPSOPHILA": "Ornamentals",
    "HARD WHEAT": "Wheat",
    "HAWTHORN": "Ornamentals",
    "HAZELNUT": "Hazelnut",
    "HAZELNUTS": "Hazelnut",
    "HEAD CABBAGE": "Cabbage",
    "HEAD LETTUCE": "Lettuce",
    "HEARTNUT": "Uncultivated Soil",
    "HEATHER": "Ornamentals",
    "HEDGE MAPLE": "Ornamentals",
    "HEMP NETTLE": "Uncultivated Soil",
    "HERB CROPS": "Vegetable",
    "HERB, SUBGROUP 19A": "Vegetable",
    "HERBACEOUS PLANTS": "Vegetable",
    "HERBS": "Vegetable",
    "HERBS AND SPICES, GROUP 19": "Vegetable",
    "HERBS, SPICES AND MEDICINAL CROPS": "Vegetable",
    "HIBISCUS": "Ornamentals",
    "HICKORY NUT": "Uncultivated Soil",
    "HIGH": "Berry Crops",
    "HIGHBUSH CRANBERRY": "Berry Crops",
    "HIMALAYABERRY": "Berry Crops",
    "HOLY THISTLE": "Uncultivated Soil",
    "HOME": "Public Health",
    "HONEYDEW MELON": "Melon",
    "HONEYSUCKLE, EDIBLE": "Berry Crops",
    "HOP": "Hops",
    "HOP CLOVER": "Clover",
    "HORSE BEAN": "Beans (Species)",
    "HORSE CHESTNUT": "Uncultivated Soil",
    "HORSERADISH": "Vegetable",
    "HOT PEPPER": "Bell Pepper",
    "HUAUZONTLE": "Vegetable",
    "HUBBARD SQUASH": "Pumpkin/Squash",
    "HUCKLEBERRY": "Berry Crops",
    "HULLBERRY": "Berry Crops",
    "HUNGARIAN MILLET": "Cereal",
    "HUSK-TOMATO": "Tomato",
    "HYACINTH": "Ornamentals",
    "HYBRID FESCUE": "Turf",
    "HYBRID TEA ROSES": "Ornamentals",
    "HYDRANGEA": "Ornamentals",
    "HYLOCEREUS UNDATUS": "Fruit",
    "HYPERICUM": "Ornamentals",
    "HYSSOP": "Vegetable",
    "ILAMA": "Fruit",
    "IMPATIENCE": "Ornamentals",
    "INDIA MUSTARD": "Brassicas",
    "INDIAN LONG PEPPER BUSH": "Vegetable",
    "INDIAN MUSTARD": "Brassicas",
    "INDIGO": "Uncultivated Soil",
    "IN-HOUSE PLANTS": "Ornamentals",
    "INSECTS, GENERAL, SPECIFY": "Public Health",
    "IRIS": "Ornamentals",
    "IRIS FAMILY": "Ornamentals",
    "ITALIAN FENNEL": "Fennel",
    "ITALIAN RYEGRASS": "Turf",
    "JABOTICABA": "Fruit",
    "JACKBEAN": "Beans (Species)",
    "JACKFRUIT": "Fruit",
    "JAPANESE APRICOT": "Apricot",
    "JAPANESE ARALIA": "Ornamentals",
    "JAPANESE BAMBOO": "Ornamentals",
    "JAPANESE BUSH CHERRY": "Cherry",
    "JAPANESE CAMELLIA": "Ornamentals",
    "JAPANESE CORNEL": "Ornamentals",
    "JAPANESE HOLLY": "Ornamentals",
    "JAPANESE HORNWORT": "Vegetable",
    "JAPANESE LAWNGRASS": "Turf",
    "JAPANESE MEDLAR": "Fruit",
    "JAPANESE MUSTARD": "Brassicas",
    "JAPANESE PEAR": "Pear",
    "JAPANESE RED CEDAR": "Forestry",
    "JAPANESE SPINDLE": "Ornamentals",
    "JAPANESE STAR ANISE": "Vegetable",
    "JAPANESE TEA": "Tea",
    "JAPANESE TURNIP": "Vegetable",
    "JAPANESE YAM": "Vegetable",
    "JAPANESE ZELKOVA": "Ornamentals",
    "JASMINE": "Ornamentals",
    "JAVA APPLE": "apple",
    "JERUSALEM ARTICHOKE": "Artichoke",
    "JERUSALEM-CHERRY": "Ornamentals",
    "JEW'S MALLOW": "Vegetable",
    "JOB'S TEARS": "Cereal",
    "JOJOBA": "Uncultivated Soil",
    "JOSTABERRY": "Berry Crops",
    "JUGLANS REGIA": "Uncultivated Soil",
    "JUJUBE": "Fruit",
    "JUJUBE, CHINESE": "Fruit",
    "JUJUBE, INDIAN": "Fruit",
    "JUNEBERRY": "Berry Crops",
    "JUNIPER": "Forestry",
    "JUNIPER, BERRY": "Berry Crops",
    "JUTE PLANT": "Uncultivated Soil",
    "JUTE, LEAVES": "Vegetable",
    "KALANCHOE": "Ornamentals",
    "KALE": "Brassicas",
    "KALE, CURLY": "Brassicas",
    "KANGKUNG": "Vegetable",
    "KENAF": "Uncultivated Soil",
    "KERSTING'S GROUNDNUT": "Peanut",
    "KHAKIWEED": "Uncultivated Soil",
    "KIDNEY BEAN": "Beans (Species)",
    "KIWI": "Kiwi",
    "KIWI FRUIT": "Kiwi",
    "KIWI PLANT": "Kiwi",
    "KIWIFRUIT": "Kiwi",
    "KIWIFRUIT, FUZZY": "Kiwi",
    "KIWIFRUIT, HARDY": "Kiwi",
    "KOHLRABI": "Brassicas",
    "KONNYAKU": "Vegetable",
    "KOREAN PINE": "Forestry",
    "KUMQUAT": "Citrus",
    "KUMQUATS": "Citrus",
    "KURRAT": "Vegetable",
    "LABLAB BEAN (HYACINTH BEAN)": "Beans (Species)",
    "LADY'S LEEK": "Leek",
    "LAMB'S LETTUCE": "Lettuce",
    "LARGE SHRUB/TREE BERRY, SUBGROUP 13-07C": "Berry Crops",
    "LARGE-ROOTED CHICORY": "Chicory/Endive",
    "LAUREL": "Ornamentals",
    "LAVACABERRY": "Berry Crops",
    "LAVENDER": "Ornamentals",
    "LAWN PENNYWORT": "Turf",
    "LAWN PLANTS": "Turf",
    "LEAF AND STEM VEGETABLE": "Vegetable",
    "LEAF BEET": "Sugarbeet",
    "LEAF LETTUCE": "Lettuce",
    "LEAF MUSTARD": "Brassicas",
    "LEAF PETIOLES, SUBGROUP 4B": "Vegetable",
    "LEAFY BRASSICA CROPS": "Brassicas",
    "LEAFY GREENS, SUBGROUP 4A": "Vegetable",
    "LEAFY VEGETABLE": "Vegetable",
    "LEAFY VEGETABLE CROPS (EXCLUDING BRASSICA)": "Vegetable",
    "LEAFY VEGETABLE GROUP": "Vegetable",
    "LEAFY VEGETABLE GROUP, GROUP 4-16": "Vegetable",
    "LEAFY VEGETABLE (INCLUDING BRASSICA LEAFY VEGETABLE)": "Vegetable",
    "LEATHERFERN": "Ornamentals",
    "LEATHER-FLOWER": "Ornamentals",
    "LEEK": "Leek",
    "LEEK, WILD": "Leek",
    "LEGUME ANIMAL FEEDS": "Vegetable",
    "LEGUME VEGETABLE CROPS": "Vegetable",
    "LEGUME VEGETABLE": "Vegetable",
    "LEGUME, FORAGE": "Vegetable",
    "LEMON": "Citrus",
    "LEMON BALM": "Vegetable",
    "LEMON/LIME, SUBGROUP 10-10B": "Citrus",
    "LEMONGRASS": "Vegetable",
    "LEMONS AND LIMES": "Citrus",
    "LENS": "Lentil",
    "LENTIL": "Lentil",
    "LENTIL (YOUNG PODS)": "Lentil",
    "LESPEDEZA": "Clover",
    "LESQUERELLA": "Rape/ Oilseed Rape",
    "LETTUCE": "Lettuce",
    "LETTUCE TREE": "Lettuce",
    "LETTUCE, BITTER": "Lettuce",
    "LETTUCE, HEAD": "Lettuce",
    "LETTUCE, LEAF": "Lettuce",
    "LETTUCE, LEAF (ROMAINE)": "Lettuce",
    "LEUCAENA": "Uncultivated Soil",
    "LICORICE": "Vegetable",
    "LILIACEAE": "Ornamentals",
    "LILIE": "Ornamentals",
    "LILIUM SP.": "Ornamentals",
    "LILY-FLOWERED TULIPS": "Ornamentals",
    "LIMA BEAN": "Beans (Species)",
    "LIMA BEAN (YOUNG PODS AND/OR IMMATURE BEANS)": "Beans (Species)",
    "LIME": "Citrus",
    "LINGONBERRY": "Berry Crops",
    "LINSEED": "Linseed/Flax",
    "LIQUORICE": "Vegetable",
    "LIQUORICE, ROOTS": "Vegetable",
    "LISIANTHUS": "Ornamentals",
    "LITCHI": "Fruit",
    "LOBESIA BOTRANA": "Public Health",
    "LOGANBERRY": "Berry Crops",
    "LONGAN": "Fruit",
    "LOQUAT": "Fruit",
    "LOTUS ROOT": "Vegetable",
    "LOVAGE": "Vegetable",
    "LOW GROWING BERRY, EXCEPT STRAWBERRY, SUBGROUP 13": "Berry Crops",
    "LOW GROWING BERRY, SUBGROUP 13": "Berry Crops",
    "LOWBERRY": "Berry Crops",
    "LOWBUSH BLUEBERRY": "Berry Crops",
    "LUCRETIABERRY": "Berry Crops",
    "LUNARIA": "Ornamentals",
    "LUPIN": "Lupin",
    "LUPINE": "Lupin",
    "LUPINUS ALBUS": "Lupin",
    "LUPINUS ANGUSTIFOLIUS": "Lupin",
    "LYCHEE": "Fruit",
    "MACADAMIA NUT": "Fruit",
    "MACADAMIA NUT (BUSH NUT)": "Fruit",
    "MACADAMIA NUTS": "Fruit",
    "MAIZE": "Corn",
    "MAKUWA MELON": "Melon",
    "MALLOW": "Vegetable",
    "MAMMEY APPLE": "Fruit",
    "MAMMOTH BLACKBERRY": "Berry Crops",
    "MANDARIN": "Citrus",
    "MANDARINS": "Citrus",
    "MANDEVILLA": "Ornamentals",
    "MANGABA": "Fruit",
    "MANGETOUT OR MANGETOUT PEA": "Pea",
    "MANGO": "Mango",
    "MANGO MELON": "Melon",
    "MANGOSTEEN": "Fruit",
    "MANIOC": "Cassava / Manioc",
    "MARIGOLD": "Ornamentals",
    "MARIGOLD FLOWERS": "Ornamentals",
    "MARIONBERRY": "Berry Crops",
    "MARJORAM": "Vegetable",
    "MARMALADEBOX": "Fruit",
    "MARROW": "Pumpkin/Squash",
    "MARROW SQUASH": "Pumpkin/Squash",
    "MARROW-STEM CABBAGE OR MARROW STEM KALE": "Brassicas",
    "MARSH-MALLOW": "Vegetable",
    "MASCARENEGRASS": "Turf",
    "MAT BEAN (GREEN PODS, MATURE, FRESH SEEDS)": "Beans (Species)",
    "MATRIMONY VINE": "Ornamentals",
    "MAYHAW": "Fruit",
    "MAYPOP": "Fruit",
    "MEADOW FESCUE": "Turf",
    "MEADOW GRASS": "Turf",
    "MEDICINAL PLANTS": "Public Health",
    "MEDICK": "Clover",
    "MEDLAR": "Fruit",
    "MELILOTUS SP.": "Clover",
    "MELON": "Melon",
    "MELON, NETTED": "Melon",
    "MELON, SUBGROUP 9A": "Melon",
    "MILFOIL": "Uncultivated Soil",
    "MILK THISTLE": "Ornamentals",
    "MILKVETCH": "Ornamentals",
    "MILKWEED": "Ornamentals",
    "MILLET": "Cereal",
    "MILLET, FOXTAIL": "Cereal",
    "MILLET, PEARL": "Cereal",
    "MILLET, PROSO": "Cereal",
    "MINT": "Vegetable",
    "MINTS": "Vegetable",
    "MIOGA": "Vegetable",
    "MIOGA GINGER": "Ginseng",
    "MIRABELLE": "Plum",
    "MISCANTHUS": "Triticale",
    "MIXED FOREST PLANTS": "Forestry",
    "MIZUNA": "Vegetable",
    "MOMBIN, PURPLE": "Fruit",
    "MONGOLIAN OAK": "Forestry",
    "MONKEY PUZZLE NUTS": "Fruit Tree",
    "MONTEREY PINE": "Forestry",
    "MORAS": "Berry Crops",
    "MOTH BEAN": "Beans (Species)",
    "MOUNTAIN CHERRY": "Cherry",
    "MULBERRIES": "Berry Crops",
    "MULBERRY": "Berry Crops",
    "MUNG BEAN": "Beans (Species)",
    "MUNG BEAN (GREEN PODS)": "Beans (Species)",
    "MUNTRIES": "Berry Crops",
    "MUSHROOM": "Mushrooms",
    "MUSHROOMS": "Mushrooms",
    "MUSKMELON": "Melon",
    "MUSKY GOURD": "Cucurbit",
    "MUSTARD": "Brassicas",
    "MUSTARD CROPS": "Brassicas",
    "MUSTARD GREENS": "Brassicas",
    "MUSTARD SPINACH": "Brassicas",
    "MUSTARD, BROWN": "Brassicas",
    "MUSTARD, SEED": "Brassicas",
    "MYRRH": "Forestry",
    "NA  PLEASE REFER TO CROP REMARK": "Uncultivated Soil",
    "NARANJILLA": "Fruit",
    "NARCISSUS": "Ornamentals",
    "NARROW-LEAVED PLANTAIN": "Ornamentals",
    "NATIVE CURRANT": "Berry Crops",
    "NAVY BEAN": "Beans (Species)",
    "NECTARBERRY": "Berry Crops",
    "NECTARINE": "Stone Fruit",
    "NETTLE": "Vegetable",
    "NEW GUINEA HYBRIDS IMPATIENS": "Ornamentals",
    "NEW ZEALAND SPINACH": "Spinach",
    "NICARAGUAN COCOA SHADE TREE": "Cocoa/Cacao",
    "NIGER SEED": "Rape/ Oilseed Rape",
    "NIGHT-BLOOMING CEREUS": "Ornamentals",
    "NO NAME": "Uncultivated Soil",
    "NOBLE CHAMOMILE": "Vegetable",
    "NON CROP": "Uncultivated Soil",
    "NON-CROP LAND": "Uncultivated Soil",
    "NON-CROP LAND": "Uncultivated Soil",
    "NORTHERN DEWBERRY": "Berry Crops",
    "NOT KNOWN": "Uncultivated Soil",
    "NURSERY STOCK": "Ornamentals",
    "NUT CROPS": "Hazelnut",
    "NUT, PINE": "Hazelnut",
    "NUT, TREE, GROUP 14": "Fruit Tree",
    "NUT, TREE, GROUP 14": "Fruit Tree",
    "OAK": "Forestry",
    "OAT": "Oat",
    "OAT (SPRING)": "Oat",
    "OATS": "Oat",
    "OIL PALM": "Oil Palm",
    "OIL RADISH": "Radish",
    "OILSEED": "Rape/ Oilseed Rape",
    "OILSEED RAPE": "Rape/ Oilseed Rape",
    "OILSEED TURNIP": "Vegetable",
    "OILSEED, GROUP 20": "Rape/ Oilseed Rape",
    "OKRA": "Curcubit",
    "OLALLIEBERRY": "Berry Crops",
    "OLIVE": "Olive",
    "OLIVE OIL": "Olive",
    "OLIVES": "Olive",
    "ONION": "Onion",
    "ONION, BELTSVILLE BUNCHING": "Onion",
    "ONION, BULB": "Onion",
    "ONION, BULB, SUBGROUP 3": "Onion",
    "ONION, CHINESE, BULB": "Onion",
    "ONION, DRY BULB": "Onion",
    "ONION, DRY BULB AND GREEN": "Onion",
    "ONION, FRESH": "Onion",
    "ONION, GREEN": "Onion",
    "ONION, GREEN, SUBGROUP 3": "Onion",
    "ONION, MACROSTEM": "Onion",
    "ONION, PEARL": "Onion",
    "ONION, POTATO, BULB": "Onion",
    "ONION, TREE, TOPS": "Onion",
    "ONION, WELSH": "Onion",
    "OPIUM POPPY": "Uncultivated Soil",
    "OPUNTIA COCHENILLIFERA": "Ornamentals",
    "ORACH": "Vegetable",
    "ORACHE": "Vegetable",
    "ORANGE": "Citrus",
    "ORANGE JESSAMINE": "Ornamentals",
    "ORANGE, SOUR": "Citrus",
    "ORANGE, SUBGROUP 10": "Citrus",
    "ORANGE, SWEET": "Citrus",
    "ORANGES, SWEET, SOUR (INCLUDING ORANGE-LIKE HYBRIDS): SEVERAL CULTIVARS": "Citrus",
    "ORCHARD": "Fruit Tree",
    "ORCHARDGRASS": "Turf",
    "ORCHID": "Ornamentals",
    "OREGANO": "Vegetable",
    "OREGON EVERGREEN BERRY": "Berry Crops",
    "ORIENTAL GARLIC": "Garlic",
    "ORNAMENTAL APPLE": "Ornamentals",
    "ORNAMENTAL BROADLEAF TREES AND SHRUBS": "Ornamentals",
    "ORNAMENTAL BULB PLANTS": "Ornamentals",
    "ORNAMENTAL CONIFER PLANTS": "Ornamentals",
    "ORNAMENTAL HERBACEOUS PERENNIAL PLANTS": "Ornamentals",
    "ORNAMENTAL HERBACEOUS PERENNIALS": "Ornamentals",
    "ORNAMENTAL HERBACEOUS PLANTS": "Ornamentals",
    "ORNAMENTAL PLANTS": "Ornamentals",
    "ORNAMENTAL SHRUBS PLANTS": "Ornamentals",
    "ORNAMENTAL WOODY PLANTS": "Ornamentals",
    "OSAGE ORANGE": "Ornamentals",
    "OXHEART CABBAGE": "Cabbage",
    "OYSTER MUSHROOM": "Mushrooms",
    "PAEONY": "Ornamentals",
    "PAK-CHOI": "Brassicas",
    "PAK-CHOI OR PAKSOI": "Brassicas",
    "PALM GRASS": "Turf",
    "PALM HEART": "Vegetable",
    "PALM HEARTS": "Vegetable",
    "PALM TREES": "Ornamentals",
    "PALM, OIL": "Oil Palm",
    "PAPAYA": "Fruit",
    "PAPRICA": "Bell Pepper",
    "PAPRIKA": "Bell Pepper",
    "PARA NUT": "Uncultivated Soil",
    "PARA RUBBER": "Rubber Tree",
    "PARSLEY": "Parsley",
    "PARSLEY, FRESH LEAVES": "Parsley",
    "PARSLEY, TURNIP ROOTED": "Parsley",
    "PARSNIP": "Vegetable",
    "PARTRIDGEBERRY": "Berry Crops",
    "PASPALUM": "Turf",
    "PASSION FRUIT": "Fruit",
    "PASSIONFLOWER": "Ornamentals",
    "PASSIONFRUIT": "Fruit",
    "PASTURE": "Turf",
    "PAWPAW": "Fruit",
    "PEA": "Pea",
    "PEA AND BEAN, DRIED SHELLED, EXCEPT SOYBEAN, SUBGROUP 6C": "Beans (Species)",
    "PEA AND BEAN, SUCCULENT SHELLED, SUBGROUP 6B": "Beans (Species)",
    "PEA EGGPLANT": "Aubergine/Eggplant",
    "PEA, DRY": "Pea",
    "PEACH": "Peach",
    "PEACH TOMATO (SYNONYM: COCONA)": "Tomato",
    "PEANUT": "Peanut",
    "PEAR": "Pear",
    "PEAR, ASIAN": "Pear",
    "PEAR, ORIENTAL": "Pear",
    "PEARL LUPINE": "Lupin",
    "PEARL ONION": "Onion",
    "PECAN": "Uncultivated Soil",
    "PEKING CABBAGE": "Cabbage",
    "PELARGONIUM HORTORUM": "Ornamentals",
    "PEPINO": "Zucchini",
    "PEPPER": "Bell Pepper",
    "PEPPER BUSH": "Ornamentals",
    "PEPPER, BLACK": "Vegetable",
    "PEPPER, BLACK; WHITE": "Vegetable",
    "PEPPER, SICHUAN": "Vegetable",
    "PEPPER/EGGPLANT, SUBGROUP 8": "Bell Pepper",
    "PEPPERGRASS": "Vegetable",
    "PEPPERMINT": "Vegetable",
    "PEPPERS": "Bell Pepper",
    "PEPPERS, CHILI": "Bell Pepper",
    "PEPPERS, SWEET (INCLUDING PIMENTO OR PIMIENTO)": "Bell Pepper",
    "PERENNIAL": "Perennial Ryegrass",
    "PERENNIAL GRASSES": "Perennial Ryegrass",
    "PERENNIAL NETTLE": "Perennial Ryegrass",
    "PERENNIAL RYEGRASS": "Perennial Ryegrass",
    "PERENNIAL SAVORY": "Vegetable",
    "PERILLA": "Vegetable",
    "PERIWINKLE": "Ornamentals",
    "PERSIAN CLOVER": "Clover",
    "PERSIAN MELON": "Melon",
    "PERSIMMON": "Fruit",
    "PERSIMMON, AMERICAN": "Fruit",
    "PERSIMMON, JAPANESE": "Fruit",
    "PETSAI": "Cabbage",
    "PHACELIA": "Uncultivated Soil",
    "PHASEOLUS": "Beans (Species)",
    "PHASEOLUS SP.": "Beans (Species)",
    "PHASEOLUS VULGARIS": "Beans (Species)",
    "PHASEY BEAN": "Beans (Species)",
    "PHENOMENALBERRY": "Berry Crops",
    "PHOENIX": "Ornamentals",
    "PHOTINIA": "Ornamentals",
    "PIGEON PEA": "Beans (Species)",
    "PIMA COTTON/JUTE": "Cotton/Jute",
    "PIMENTO": "Bell Pepper",
    "PINE": "Forestry",
    "PINEAPPLE": "Berry crops",
    "PINEAPPLE MELON": "Melon",
    "PINOPSIDA": "Forestry",
    "PINTO BEAN": "Beans (Species)",
    "PINUS TAEDA": "Forestry",
    "PISTACHIO": "Peanut",
    "PISTACHIO NUT": "Peanut",
    "PISUM SATIVUM": "Pea",
    "PITAYA": "Fruit",
    "PLANE": "Forestry",
    "PLANTAE": "Uncultivated Soil",
    "PLANTAIN": "Vegetable",
    "PLANTAIN LEAVES": "Vegetable",
    "PLANTAIN, BUCKTHORN": "Vegetable",
    "PLANTES EN POT": "Ornamentals",
    "PLANTS IN GRASSLAND": "Turf",
    "PLANTS IN MEADOWS": "Turf",
    "PLANTS IN PASTURES": "Turf",
    "PLATANUS": "Forestry",
    "PLUM": "Plum",
    "PLUM APRICOT": "Apricot",
    "PLUM, DAMSON": "Plum",
    "PLUM, JAPANESE": "Plum",
    "PLUM, OR PRUNE PLUM, SUBGROUP 12": "Plum",
    "PLUMCOT": "Stone Fruit",
    "POA PRATENSIS": "Turf",
    "PODOACEAE": "Turf",
    "POINSETTIA": "Ornamentals",
    "POME FRUITS": "Pome Fruit",
    "POMEGRANATE": "Deciduous Fruit",
    "POMELO": "Citrus",
    "POMPON TREE": "Ornamentals",
    "POPCORN": "Corn",
    "POPLAR": "Forestry",
    "POPPY": "Uncultivated Soil",
    "POT MARIGOLD": "Ornamentals",
    "POT PLANTS": "Ornamentals",
    "POTATO": "Potato",
    "POTATO YAM": "Potato",
    "POTENTILLA ERECTA": "Ornamentals",
    "POTENTILLA FREYNIANA": "Ornamentals",
    "PRICKLY CYCAD": "Ornamentals",
    "PRICKLY PEAR": "Fruit",
    "PRIMROSE": "Ornamentals",
    "PRIMROSE, ENGLISH": "Ornamentals",
    "PROSO MILLET": "Cereal",
    "PRUNUS (ORNAMENTAL SPECIES)": "Ornamentals",
    "PRUNUS LAUROCERASUS": "Ornamentals",
    "PSEUDOSTELLARIA HETEROPHYLLA": "Vegetable",
    "PUBLIC HEALTH": "Public Health",
    "PULASAN": "Fruit",
    "PULSES": "Beans (Species)",
    "PUMMELO": "Citrus",
    "PUMPKIN": "Pumpkin/Squash",
    "PUMPKIN (HYBRIDS)": "Pumpkin/Squash",
    "PUMPKINS": "Pumpkin/Squash",
    "PURPLE CLOVER": "Clover",
    "PURPLE VEIN ROCKET": "Ornamentals",
    "PURSLANE": "Vegetable",
    "PURSLANE, GARDEN": "Vegetable",
    "PURSLANE, WINTER": "Vegetable",
    "PYRETHRUM": "Uncultivated Soil",
    "QUACKGRASS": "Turf",
    "QUEENSLAND BLUE GRASS": "Turf",
    "QUINCE": "Fruit",
    "QUINOA": "Cereal",
    "RADICCHIO (RED CHICORY)": "Chicory/Endive",
    "RADISH": "Radish",
    "RADISH, JAPANESE": "Radish",
    "RADISH, LEAVES": "Radish",
    "RADISH, ORIENTAL (DAIKON)": "Radish",
    "RAILS": "Uncultivated Soil",
    "RAKKYO": "Vegetable",
    "RAMBUTAN": "Fruit",
    "RANGEBERRY": "Berry Crops",
    "RAPE": "Rape/ Oilseed Rape",
    "RAPE GREENS": "Vegetable",
    "RAPESEED": "Rape/ Oilseed Rape",
    "RAPESEED, SUBGROUP 20A": "Rape/ Oilseed Rape",
    "RAPHANUS SATIVUS": "Radish",
    "RAPINI": "Vegetable",
    "RASPBERRIES, RED, BLACK": "Berry Crops",
    "RASPBERRY": "Berry Crops",
    "RASPBERRY, BLACK AND RED": "Berry Crops",
    "RATOON SUGARCANE": "Sugarcane",
    "RATTLESNAKE BELOPERONE": "Ornamentals",
    "RAVENBERRY": "Berry Crops",
    "RED CABBAGE": "Cabbage",
    "RED CLOVER": "Clover",
    "RED CURRANT": "Berry Crops",
    "RED FESCUE": "Turf",
    "RED GRAM": "Beans (Species)",
    "RED PEPPER": "Bell Pepper",
    "RED RASPBERRY": "Berry Crops",
    "REHMANN LIPPIA": "Ornamentals",
    "RHODODENDRON": "Ornamentals",
    "RHUBARB": "Vegetable",
    "RIBES": "Berry Crops",
    "RIBES SP.": "Berry Crops",
    "RIBWORT PLANTAIN": "Uncultivated Soil",
    "RICE": "Rice",
    "RICE BEAN": "Beans (Species)",
    "RICE, ROUGH": "Rice",
    "ROCKET": "Lettuce",
    "ROMAINE LETTUCE, COS LETTUCE": "Lettuce",
    "ROOT AND TUBER VEGETABLE": "Vegetable",
    "ROOT CROPS": "Vegetable",
    "ROOT PARSLEY": "Parsley",
    "ROOT VEGETABLE PLANTS": "Vegetable",
    "ROOT VEGETABLE": "Vegetable",
    "ROOTED CELERY": "Celery",
    "ROSE": "Ornamentals",
    "ROSE APPLE": "Fruit",
    "ROSE HIP": "Berry Crops",
    "ROSE OF SHARON": "Ornamentals",
    "ROSELLE": "Vegetable",
    "ROSEMARY": "Vegetable",
    "ROSSBERRY": "Berry Crops",
    "ROWAN": "Berry Crops",
    "ROYAL AZALEA": "Ornamentals",
    "RUBBER TREE": "Rubber Tree",
    "RUBUS": "Berry Crops",
    "RUCOLA": "Vegetable",
    "RUE": "Vegetable",
    "RUNNER BEAN": "Beans (Species)",
    "RUTABAGA": "Vegetable",
    "RYE": "Rye",
    "RYEGRASS": "Turf",
    "RYEGRASS, PERENNIAL": "Perennial Ryegrass",
    "SACCHARUM OFFICINARUM": "Sugarcane",
    "SAFFLOWER": "Rape/ Oilseed Rape",
    "SAFFLOWER SEED": "Rape/ Oilseed Rape",
    "SAGE": "Vegetable",
    "SAGE AND RELATED SALVIA SPECIES": "Vegetable",
    "SAGEBRUSH": "Uncultivated Soil",
    "SAINFOIN": "Uncultivated Soil",
    "SALAD": "Vegetable",
    "SALAD BURNET": "Lettuce",
    "SALAD ROCKET": "Lettuce",
    "SALAL": "Berry Crops",
    "SALIX": "Forestry",
    "SALSIFY": "Vegetable",
    "SALVIA": "Vegetable",
    "SANTA CLAUS MELON": "Melon",
    "SAPODILLA": "Fruit",
    "SAPOTE, BLACK": "Fruit",
    "SAPOTE, MAMEY": "Fruit",
    "SAPOTE, WHITE": "Fruit",
    "SATSUKI AZALEA": "Ornamentals",
    "SATSUMA MANDARIN": "Citrus",
    "SATSUMA OR SATSUMA MANDARIN": "Citrus",
    "SAVORY, SUMMER": "Vegetable",
    "SAVOY CABBAGE": "Cabbage",
    "SCALLOP SQUASH": "Pumpkin/Squash",
    "SCARLET EGGPLANT": "Aubergine/Eggplant",
    "SCARLET RUNNER BEAN": "Beans (Species)",
    "SCHISANDRA BERRY": "Berry Crops",
    "SCHIZOLOBIUM PARAHYBA": "Forestry",
    "SCORZONERA": "Vegetable",
    "SCOTS PINE": "Forestry",
    "SEA BUCKTHORN": "Berry Crops",
    "SEA": "Uncultivated Soil",
    "SEED": "Uncultivated Soil",
    "SEEDED UPLAND RICE": "Rice",
    "SEPTORIA CANNABIS": "Public Health",
    "SERRADELLA": "Uncultivated Soil",
    "SESAME": "Rape/ Oilseed Rape",
    "SESAME SEED": "Rape/ Oilseed Rape",
    "SHADDOCKS OR POMELOS": "Citrus",
    "SHALLOT": "Onion",
    "SHALLOT, BULB": "Onion",
    "SHALLOT, FRESH LEAVES": "Onion",
    "SHAWNEE BLACKBERRY": "Berry Crops",
    "SHEEP'S FESCUE": "Turf",
    "SHELL FRUITS": "Uncultivated Soil",
    "SHEPHERD'S PURSE": "Vegetable",
    "SHRUB ROSES": "Ornamentals",
    "SHRUBBY PLANTS": "Ornamentals",
    "SIBERIAN GINSENG": "Ginseng",
    "SIBERIAN PEA TREE": "Ornamentals",
    "SIBERIAN WALLFLOWER": "Ornamentals",
    "SIGNALGRASS": "Turf",
    "SILVERGRASS": "Turf",
    "SILVER-TOP EUCALYPTUS": "Forestry",
    "SILYBUM MARIANUM": "Uncultivated Soil",
    "SIX-ROWED BARLEY": "Barley",
    "SKUNK CURRANT": "Berry Crops",
    "SLOE": "Plum",
    "SMALL CRANBERRY": "Berry Crops",
    "SMALL FRUIT VINE CLIMBING, EXCEPT FUZZY KIWIFRUIT SUBGROUP 13": "Berry Crops",
    "SMALL FRUIT VINE CLIMBING, EXCEPT GRAPE, SUBGROUP 13": "Berry Crops",
    "SMALL FRUIT VINE CLIMBING, SUBGROUP 13": "Berry Crops",
    "SMALL RADISH": "Radish",
    "SNAKE MELON": "Melon",
    "SNAP BEAN": "Beans (Species)",
    "SNAPDRAGON": "Ornamentals",
    "SNOW PEA": "Pea",
    "SOFT BROME": "Turf",
    "SOFT WHEAT": "Wheat",
    "SOIL": "Uncultivated Soil",
    "SOLANACEAE": "Vegetable",
    "SORGHUM": "Sorghum",
    "SORGHUM BICOLOR": "Sorghum",
    "SORGHUM, FORAGE": "Sorghum",
    "SORGHUM, GRAIN": "Sorghum",
    "SORGHUM, GRAIN, BRAN": "Sorghum",
    "SORGHUM, GRAIN, FORAGE": "Sorghum",
    "SORGHUM, SWEET": "Sorghum",
    "SORGO OR SORGHUM, SWEET": "Sorghum",
    "SORREL": "Vegetable",
    "SORREL, COMMON, AND RELATED RUMEX SPECIES": "Vegetable",
    "SOUR CHERRY": "Cherry",
    "SOUR DOCK": "Vegetable",
    "SOURSOP": "Fruit",
    "SOUTHERN DEWBERRY": "Berry Crops",
    "SOUTHERN NAIAD": "Uncultivated Soil",
    "SOUTHERN PEA": "Pea",
    "SOYBEAN": "Soybean",
    "SOYBEAN (IMMATURE SEED)": "Soybean",
    "SOYBEAN, VEGETABLE": "Soybean",
    "SPAGHETTI SQUASH": "Pumpkin/Squash",
    "SPANISH LIME": "Fruit",
    "SPEARMINT": "Vegetable",
    "SPELT": "Wheat",
    "SPELT (SPRING)": "Wheat",
    "SPELT (WINTER)": "Wheat",
    "SPICE CROPS": "Vegetable",
    "SPICE PLANTS": "Vegetable",
    "SPICE, SUBGROUP 19B": "Vegetable",
    "SPICES": "Vegetable",
    "SPIKENARD": "Vegetable",
    "SPINACH": "Spinach",
    "SPINACH BEET": "Sugarbeet",
    "SPINACH MUSTARD": "Brassicas",
    "SPINACH TREE": "Vegetable",
    "SPINACH, NEW ZEALAND": "Spinach",
    "SPINACH, VINE (MALABAR SPINACH, INDIAN SPINACH)": "Spinach",
    "SPINDLE": "Ornamentals",
    "SPONGE TREE": "Ornamentals",
    "SPORTS TURF": "Turf",
    "SPRING BARLEY": "Barley",
    "SPRING CEREALS": "Cereal",
    "SPRING CROCUS": "Ornamentals",
    "SPRING DURUM WHEAT": "Wheat",
    "SPRING OAT": "Oat",
    "SPRING ONION": "Onion",
    "SPRING RAPE": "Rape/ Oilseed Rape",
    "SPRING RYE": "Rye",
    "SPRING TRITICALE": "Triticale",
    "SPRING TURNIP RAPE": "Rape/ Oilseed Rape",
    "SPRING WHEAT": "Wheat",
    "SPRUCE": "Forestry",
    "SQUASH": "Pumpkin/Squash",
    "SQUASH, SUMMER": "Pumpkin/Squash",
    "SQUASH, WINTER": "Pumpkin/Squash",
    "SQUASH/CUCUMBER, SUBGROUP 9B": "Cucurbit",
    "ST.JOHNSWORT": "Uncultivated Soil",
    "STACHYS SP.": "Vegetable",
    "STAR APPLE": "Fruit",
    "STARFRUIT": "Fruit",
    "STEPHANOTIS": "Ornamentals",
    "STEVIA": "Vegetable",
    "STOKES ASTER": "Ornamentals",
    "STONE FRUITS": "Stone Fruit",
    "STONE LEEK": "Leek",
    "STORED PRODUCT": "Uncultivated Soil",
    "STRAWBERRIES, WILD": "Strawberries",
    "STRAWBERRY": "Strawberries",
    "STRAWBERRY BUSH": "Ornamentals",
    "STUBBLE (CEREAL)": "Cereal",
    "SUB": "Vegetable",
    "SUB": "Vegetable",
    "SUB": "Brassicas",
    "SUCCORY": "Chicory/Endive",
    "SUDANGRASS": "Sorghum",
    "SUGAR APPLE": "Fruit",
    "SUGAR BEET": "Sugarbeet",
    "SUGAR CANE": "Sugarcane",
    "SUGAR MAIZE": "Corn",
    "SUGAR PEA": "Pea",
    "SUGAR SNAP PEA": "Pea",
    "SUGARCANE": "Sugarcane",
    "SUMMER SAVORY": "Vegetable",
    "SUMMER SQUASH": "Pumpkin/Squash",
    "SUMMER-FLOWERING ORNAMENTAL PLANTS": "Ornamentals",
    "SUNFLOWER": "Sunflower",
    "SUNFLOWER, SUBGROUP 20B": "Sunflower",
    "SURINAM CHERRY": "Fruit",
    "SWALLOWWORT GENTIAN": "Uncultivated Soil",
    "SWEDE": "Vegetable",
    "SWEET BASIL": "Vegetable",
    "SWEET CHAMOMILE": "Vegetable",
    "SWEET CHERRY": "Cherry",
    "SWEET CHESTNUT": "Uncultivated Soil",
    "SWEET CORN": "Corn",
    "SWEET CORN (CORN": "Corn",
    "SWEET FENNEL": "Fennel",
    "SWEET LEMON": "Citrus",
    "SWEET LIME": "Citrus",
    "SWEET LUPIN": "Lupin",
    "SWEET MAIZE": "Corn",
    "SWEET ORANGE": "Citrus",
    "SWEET PEA": "Pea",
    "SWEET PEPPER": "Bell Pepper",
    "POTATO": "Potato",
    "SWEET ROCKET": "Lettuce",
    "SWEET-BAY": "Vegetable",
    "SWIETENIA MACROPHYLLA": "Forestry",
    "SWISS CHARD": "Lettuce",
    "SYRIAN PLUM": "Plum",
    "TALL FESCUE": "Turf",
    "TAMARIND": "Fruit",
    "TANGELO": "Citrus",
    "TANGELO, SMALL AND MEDIUM SIZED CULTIVARS": "Citrus",
    "TANGERINE": "Citrus",
    "TARA VINE": "Ornamentals",
    "TARO": "Vegetable",
    "TARRAGON": "Vegetable",
    "TAYBERRY": "Berry Crops",
    "TEA": "Tea",
    "TEA PLANTS": "Tea",
    "TEA TREE": "Ornamentals",
    "TEA, GREEN, BLACK (BLACK, FERMENTED AND DRIED)": "Tea",
    "TEAS": "Tea",
    "TEAS (TEA AND HERB TEAS)": "Tea",
    "TECTONA GRANDIS": "Forestry",
    "TEFF OR TEF": "Cereal",
    "TEOSINTE": "Corn",
    "TEPARY BEAN": "Beans (Species)",
    "TERRACES AND WAYS": "Uncultivated Soil",
    "THEOBROMA CACAO": "Cocoa/Cacao",
    "THISTLE": "Uncultivated Soil",
    "THYME": "Vegetable",
    "TIFTON BURCLOVER": "Clover",
    "TIMBER STAND AS FOREST PLANTS": "Forestry",
    "TIMOTHY": "Turf",
    "TOBACCO": "Tobacco",
    "TOMATILLO": "Tomato",
    "TOMATO": "Tomato",
    "TOMATO, SUBGROUP 8": "Tomato",
    "TOTAL WEED CONTROL": "Uncultivated Soil",
    "TRAGOPOGON SP.": "Vegetable",
    "TRANSPLANTED ONION": "Onion",
    "TRANSPLANTED PADDY RICE": "Rice",
    "TREE NURSERIES": "Ornamentals",
    "TREE NUTS": "Fruit Tree",
    "TREE ONION": "Onion",
    "TREE TOMATO": "Tomato",
    "TREES": "Forestry",
    "TREFOIL": "Clover",
    "TRIDENT MAPLE": "Forestry",
    "TRIFOLIUM": "Clover",
    "TRIGONELLA FOENUM-GRAECUM": "Vegetable",
    "TRITICALE": "Triticale",
    "TRITICUM SP.": "Wheat",
    "TRITORDEUM SP.": "Wheat",
    "TROPICAL AND SUBTROPICAL FRUIT, INEDIBLE PEEL, GROUP 24": "Fruit",
    "TROPICAL AND SUBTROPICAL, MEDIUM TO LARGE FRUIT, ROUGH OR HAIRY, INEDIBLE PEEL SUBGROUP 24C": "Fruit",
    "TROPICAL AND SUBTROPICAL, MEDIUM TO LARGE FRUIT, SMOOTH, INEDIBLE PEEL SUBGROUP 24B": "Fruit",
    "TROPICAL AND SUBTROPICAL, SMALL FRUIT, INEDIBLE PEEL SUBGROUP 24A": "Fruit",
    "TROPICAL FRUIT PLANTS": "Fruit",
    "TRUE ALOE": "Ornamentals",
    "TRUE CANTALOUPE": "Melon",
    "TUBEROUS AND CORM VEGETABLE": "Vegetable",
    "TUBEROUS PLANTS": "Vegetable",
    "TULIP": "Ornamentals",
    "TUNG-OIL TREE": "Uncultivated Soil",
    "TURBAN SQUASH": "Pumpkin/Squash",
    "TURF": "Turf",
    "TURF PLANTS": "Turf",
    "TURMERIC": "Vegetable",
    "TURMERIC, ROOT": "Vegetable",
    "TURNIP": "Vegetable",
    "TURNIP CABBAGE": "Vegetable",
    "TURNIP GREENS": "Vegetable",
    "TURNIP LEAVES OR TOPS": "Vegetable",
    "TURNIP TOPS": "Vegetable",
    "TURNIP-ROOTED CHERVIL": "Vegetable",
    "TURNIP-ROOTED PARSLEY": "Parsley",
    "TUSSOCK BELLFLOWER": "Ornamentals",
    "TWO-ROWED BARLEY": "Barley",
    "TYFON": "Vegetable",
    "TYFON, LEAVES": "Vegetable",
    "UDO": "Vegetable",
    "UMBELLIFERAE": "Vegetable",
    "URD BEAN": "Beans (Species)",
    "USE IN AGRICULTURE": "Uncultivated Soil",
    "VALERIAN": "Vegetable",
    "VEGETABLE CROPS": "Vegetable",
    "VEGETABLE FERN": "Vegetable",
    "VEGETABLE PLANTS": "Vegetable",
    "VEGETABLE, BRASSICA, LEAFY, GROUP 5": "Brassicas",
    "VEGETABLE, BULB, GROUP 3": "Vegetable",
    "VEGETABLE, BULB, GROUP 3-7": "Vegetable",
    "VEGETABLE, CUCURBIT, GROUP 9": "Cucurbit",
    "VEGETABLE, FOLIAGE OF LEGUME, EXCEPT SOYBEAN, SUBGROUP 7A": "Vegetable",
    "VEGETABLE, FRUITING, GROUP 8": "Vegetable",
    "VEGETABLE, FRUITING, GROUP 8": "Vegetable",
    "VEGETABLE, LEAFY, EXCEPT BRASSICA, GROUP 4": "Vegetable",
    "VEGETABLE, LEAVES OF ROOT AND TUBER, GROUP 2": "Vegetable",
    "VEGETABLE, LEGUME, EDIBLE PODDED, SUBGROUP 6A": "Vegetable",
    "VEGETABLE, LEGUME, GROUP 6": "Vegetable",
    "VEGETABLE, ROOT AND TUBER, GROUP 1": "Vegetable",
    "VEGETABLE, ROOT, EXCEPT SUGARBEET, SUBGROUP 1B": "Vegetable",
    "VEGETABLE, ROOT, SUBGROUP 1A": "Vegetable",
    "VEGETABLE, STALK AND STEM, SUBGROUP 22A": "Vegetable",
    "VEGETABLE, TUBEROUS AND CORM, EXCEPT POTATO, SUBGROUP 1D": "Vegetable",
    "VEGETABLE, TUBEROUS AND CORM, SUBGROUP 1C": "Vegetable",
    "VEGETABLE": "Vegetable",
    "VEGETABLE, CUCURBITS": "Cucurbit",
    "VEGETABLE, ONION, BULB AND ONION, GREEN, GROUP 3-7": "Onion",
    "VERBENA HYBRIDS": "Ornamentals",
    "VERBENA LITORALIS": "Ornamentals",
    "VERONICA": "Ornamentals",
    "VETCH": "Uncultivated Soil",
    "VETCH, CROWN": "Uncultivated Soil",
    "VIBURNUM": "Ornamentals",
    "VICIA FABA": "Beans (Species)",
    "VICIA SP.": "Beans (Species)",
    "VIOLET": "Ornamentals",
    "VIRGINIAN TOBACCO": "Tobacco",
    "WALNUT": "Hazelnut",
    "WALNUT, BLACK": "Hazelnut",
    "WALNUT, ENGLISH (PERSIAN)": "Hazelnut",
    "WALNUTS": "Hazelnut",
    "WASABI": "Ginseng",
    "WATER": "Uncultivated Soil",
    "WATER BAMBOO": "Vegetable",
    "WATER CALTROP": "Vegetable",
    "WATER DROPWORT": "Vegetable",
    "WATER SPINACH": "Spinach",
    "WATERCRESS": "Vegetable",
    "WATERMELON": "Watermelon",
    "WATERWEED": "Uncultivated Soil",
    "WAX APPLE": "Fruit",
    "WAX BEAN": "Beans (Species)",
    "WAX GOURD": "Cucurbit",
    "WAX JAMBU": "Fruit",
    "WAYS AND PLACES": "Uncultivated Soil",
    "WAYS AND PLACES WITH TREES": "Forestry",
    "WEEPING WILLOW": "Forestry",
    "WELSH ONION": "Onion",
    "WEST INDIAN LEMON GRASS": "Vegetable",
    "WET-SEEDED PADDY RICE": "Rice",
    "WHEAT": "Wheat",
    "WHEAT, GRAIN": "Wheat",
    "WHITE CABBAGE": "Cabbage",
    "WHITE CLOVER": "Clover",
    "WHITE LUPIN": "Lupin",
    "WHITE LUPINE": "Lupin",
    "WHITE MUSTARD": "Brassicas",
    "WHITE SWEET LUPIN": "Lupin",
    "WHITE WILLOW": "Forestry",
    "WILD CABBAGE": "Cabbage",
    "WILD CHAMOMILE": "Vegetable",
    "WILD CHICORY": "Chicory/Endive",
    "WILD COFFEE": "Coffee",
    "WILD JUJUBE": "Fruit",
    "WILD LEEK": "Leek",
    "WILD MARJORAM": "Vegetable",
    "WILD MARJORAM (OREGANO)": "Vegetable",
    "WILD PANSY": "Ornamentals",
    "WILD PINEAPLE": "Pineapple",
    "WILD RASPBERRY": "Berry Crops",
    "WILD RICE": "Rice",
    "WILD STRAWBERRY": "Strawberries",
    "WILD TOBACCO": "Tobacco",
    "WILD TURNIP": "Vegetable",
    "WILLOW": "Forestry",
    "WILLOWHERB": "Uncultivated Soil",
    "WILLOW-LEAF EUCALYPTUS": "Forestry",
    "WINTER BARLEY": "Barley",
    "WINTER CEREALS": "Cereal",
    "WINTER CRESS, COMMON; AMERICAN": "Vegetable",
    "WINTER ENDIVE": "Chicory/Endive",
    "WINTER OAT": "Oat",
    "WINTER OILSEED RAPE": "Rape/ Oilseed Rape",
    "WINTER PURSLANE": "Vegetable",
    "WINTER RAPE": "Rape/ Oilseed Rape",
    "WINTER RYE": "Rye",
    "WINTER TRITICALE": "Triticale",
    "WINTER TURNIP RAPE": "Rape/ Oilseed Rape",
    "WINTER WHEAT": "Wheat",
    "WITLOOF": "Chicory/Endive",
    "WITLOOF CHICORY": "Chicory/Endive",
    "WITLOOF CHICORY (SPROUTS)": "Chicory/Endive",
    "WOOD": "Forestry",
    "WOODY PLANTS": "Forestry",
    "WOOLLY FOXGLOVE": "Ornamentals",
    "WORMWOOD": "Uncultivated Soil",
    "WRINKLED PEA": "Pea",
    "YACON": "Vegetable",
    "YAM": "Vegetable",
    "YAM BEAN": "Vegetable",
    "YAM, TRUE": "Vegetable",
    "YAMS": "Vegetable",
    "YARD-LONG BEAN (PODS)": "Beans (Species)",
    "YARROW": "Uncultivated Soil",
    "YELLOW LUPINE": "Lupin",
    "YELLOW-BERRIED NIGHTSHADE": "Vegetable",
    "YELLOWHORN": "Uncultivated Soil",
    "YOUNGBERRY": "Berry Crops",
    "ZARZAMORA": "Berry Crops",
    "ZEA MAYS": "Corn",
    "ZINNIA": "Ornamentals",
    "ZONAL PELARGONIUM": "Ornamentals",
    "ZUCCHINI": "Pumpkin/Squash",
    "ZUIKI": "Vegetable",
}


def map_crop_equivalent(crop_name):
    """Map a Target Crop to its Crop Equivalent (Reliance #43922)."""
    if pd.isna(crop_name) or not crop_name:
        return "Unknown"
    crop_upper = str(crop_name).strip().upper()
    result = CROP_EQUIVALENTS.get(crop_upper)
    if result:
        return result
    # Try partial match (crop might have extra spaces or slight differences)
    for key, val in CROP_EQUIVALENTS.items():
        if key in crop_upper or crop_upper in key:
            return val
    return str(crop_name).strip()  # Return original if no mapping found




# ==============================================================================
#  FIXED BUSINESS RULES - DATA PROCESSING
# ==============================================================================

def process_gap_data(df):
    """
    Apply fixed business rules to GAP data (Reliance #43988 & #43922).
    Returns ALL valid records (not aggregated) for user RIL selection.
    """
    stats = {}
    stats["total_raw"] = len(df)

    # --- RULE 1: Filter GAP Lifecycle State = "Approved" ---
    lifecycle_col = None
    # Try exact known names first
    for col in ["GAP Lifecycle State", "GAP UI State", "Agency Dossier Lifecycle State"]:
        if col in df.columns:
            lifecycle_col = col
            break
    # Try fuzzy match
    if lifecycle_col is None:
        for col in df.columns:
            col_lower = col.lower()
            if ("lifecycle" in col_lower or "ui state" in col_lower) and "state" in col_lower:
                lifecycle_col = col
                break

    if lifecycle_col:
        df_filtered = df[df[lifecycle_col].astype(str).str.strip() == "Approved"].copy()
    else:
        df_filtered = df.copy()
        st.warning("GAP Lifecycle State column not found. Using all rows.")

    stats["after_approved_filter"] = len(df_filtered)

    
    # --- Identify key columns ---
    col_mapping = {}
    target_cols = {
        "country": ["Country", "AP Country", "GAP Country Name"],
        "gap_variant": ["GAP Variant"],
        "target_crop": ["Target Crop"],
        "target_crop_code": ["Target Crop Code"],
        "max_rate": ["Max", "Product Max Rate"],
        "rate_uom": ["UoM", "Product Rate UoM"],
        "max_app": ["Max. App. Number / Season", "Max. Application Number / Season",
                    "Max App Number / Season"],
        "app_method": ["Application Method"],
        "gap_id": ["GAP UI ID", "GAP Usage Information ID"],
        "gap_remarks": ["GAP Remarks"],
        "gap_info_type": ["GAP Usage Information Type"],
    }

    for key, possible_names in target_cols.items():
        for name in possible_names:
            if name in df_filtered.columns:
                col_mapping[key] = name
                break
        if key not in col_mapping:
            for col in df_filtered.columns:
                for name in possible_names:
                    if name.lower().replace(" ", "").replace(".", "") in col.lower().replace(" ", "").replace(".", ""):
                        col_mapping[key] = col
                        break
                if key in col_mapping:
                    break

    # Check required columns
    required = ["gap_variant", "max_rate", "rate_uom"]
    missing = [k for k in required if k not in col_mapping]
    if missing:
        st.error(f"Required columns not found: {missing}. "
                 f"Available columns: {list(df_filtered.columns[:30])}")
        return None, stats

    # --- Select columns (keep all found) ---
    select_cols = {v: k for k, v in col_mapping.items()}
    df_work = df_filtered[list(col_mapping.values())].copy()
    df_work.columns = [select_cols.get(c, c) for c in df_work.columns]

    # --- RULE 2: Convert Max Rate to numeric, reject invalid ---
    df_work["max_rate_numeric"] = pd.to_numeric(df_work["max_rate"], errors="coerce")
    invalid_rate = ["0", "-", ""]
    df_work = df_work[
        (df_work["max_rate_numeric"].notna()) &
        (df_work["max_rate_numeric"] > 0) &
        (~df_work["max_rate"].astype(str).str.strip().isin(invalid_rate))
    ].copy()
    stats["after_rate_filter"] = len(df_work)

    # --- RULE 3: Max Applications - if blank/zero/non-numeric -> assume 1 ---
    if "max_app" in df_work.columns:
        df_work["max_app_numeric"] = pd.to_numeric(df_work["max_app"], errors="coerce")
        df_work["max_app_numeric"] = df_work["max_app_numeric"].apply(
            lambda x: 1 if (pd.isna(x) or x <= 0) else x
        )
    else:
        df_work["max_app_numeric"] = 1

    # --- RULE 4: Filter invalid UoM ---
    invalid_uom = ["0", "-", ""]
    df_work = df_work[
        (~df_work["rate_uom"].astype(str).str.strip().isin(invalid_uom)) &
        (df_work["rate_uom"].notna())
    ].copy()
    stats["after_validity_filter"] = len(df_work)

    # --- RULE 5: Convert to g/ha (Reliance #43988) ---
    df_work["rate_gha"] = df_work.apply(
        lambda row: convert_to_gha(row["max_rate_numeric"], row["rate_uom"]),
        axis=1
    )
    df_work = df_work[df_work["rate_gha"].notna()].copy()
    stats["after_conversion"] = len(df_work)

    # --- RULE 6: Reject if rate > 20,000 g/ha ---
    df_work = df_work[df_work["rate_gha"] <= 20000].copy()
    stats["after_max_rate_filter"] = len(df_work)

    # --- RULE 7: Calculate Season Total Rate ---
    df_work["season_total_rate"] = df_work["rate_gha"] * df_work["max_app_numeric"]

    # --- RULE 8: Remove rows with Season Total Rate >= 1,000,000 g/ha ---
    df_work = df_work[df_work["season_total_rate"] < 1000000].copy()
    stats["after_outlier_filter"] = len(df_work)
    stats["final_records"] = len(df_work)

    # --- RULE 9: Map Target Crop to Crop Equivalent (Reliance #43922) ---
    if "target_crop" in df_work.columns:
        df_work["crop_equivalent"] = df_work["target_crop"].apply(map_crop_equivalent)

    # --- Round calculated values ---
    df_work["rate_gha"] = df_work["rate_gha"].round(4)
    df_work["season_total_rate"] = df_work["season_total_rate"].round(4)

    # Stats
    if "gap_variant" in df_work.columns:
        stats["gap_variants"] = df_work["gap_variant"].nunique()
    if "country" in df_work.columns:
        stats["countries"] = df_work["country"].nunique()

    return df_work, stats


# ==============================================================================
#  EXCEL EXPORT
# ==============================================================================

def build_report_excel(report_df, stats, selected_variant=None, selected_country=None):
    """Build formatted Excel report."""
    h_fill = PatternFill("solid", fgColor="007A3D")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    d_font = Font(name="Arial", size=10)
    c_aln  = Alignment(horizontal="center", vertical="center")
    bdr    = Border(left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"))

    def write_sheet(ws, df, title):
        ws.title = title
        cols = list(df.columns)
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(1, ci, c)
            cell.fill, cell.font, cell.alignment, cell.border = h_fill, h_font, c_aln, bdr
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.font, cell.border = d_font, bdr
                cell.alignment = Alignment(vertical="center")
        for ci, c in enumerate(cols, 1):
            ml = max(len(str(c)),
                     *[len(str(ws.cell(r, ci).value or ""))
                       for r in range(2, min(ws.max_row + 1, 52))])
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = openpyxl.Workbook()

    # Main report sheet
    export_cols = [c for c in ["gap_variant", "country", "target_crop", "crop_equivalent",
                                "app_method", "max_rate", "rate_uom", "rate_gha",
                                "max_app_numeric", "season_total_rate", "gap_id", "gap_remarks"]
                   if c in report_df.columns]
    export_df = report_df[export_cols].copy()
    export_df.columns = ["GAP Variant", "Country", "Target Crop", "Crop Equivalent",
                         "Application Method", "Product Max Rate", "Rate UoM",
                         "Rate (g/ha)", "N. Applications", "Season Total (g/ha)",
                         "GAP ID", "GAP Remarks"][:len(export_cols)]
    write_sheet(wb.active, export_df.head(100000), "All Valid Data")

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    summary_rows = [
        ("RIRA Report - Summary", ""),
        ("", ""),
        ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("GAP Variant Filter", selected_variant or "All"),
        ("Country Filter", selected_country or "All"),
        ("Total Records in Report", len(report_df)),
        ("", ""),
        ("Business Rules Applied (Fixed)", ""),
        ("  1. GAP Lifecycle State = Approved", ""),
        ("  2. Product Max Rate valid (not 0/-)", ""),
        ("  3. N. Applications: if blank/zero/non-numeric = 1", ""),
        ("  4. Product Rate UoM valid", ""),
        ("  5. Convert to g/ha (Reliance #43988)", ""),
        ("  6. Reject if Rate > 20,000 g/ha", ""),
        ("  7. Season Total = Rate x N. Apps", ""),
        ("  8. Reject if Season Total >= 1,000,000 g/ha", ""),
        ("  9. Map Target Crop (Reliance #43922)", ""),
    ]
    for ri, (k, v) in enumerate(summary_rows, 1):
        ka, va = ws.cell(ri, 1, k), ws.cell(ri, 2, v)
        ka.font = Font(name="Arial", size=10, bold=(ri == 1 or ri >= 8))
        va.font = Font(name="Arial", size=10)
        if ri == 1:
            ka.fill = PatternFill("solid", fgColor="007A3D")
            ka.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==============================================================================
#  SIDEBAR
# ==============================================================================

with st.sidebar:
    st.markdown('<div class="section-title">Business Rules (Fixed)</div>',
                unsafe_allow_html=True)
    st.markdown("""
    <div class="rule-box filter">
    <strong>Rule 1:</strong> GAP Lifecycle State = "Approved"
    </div>
    <div class="rule-box filter">
    <strong>Rule 2:</strong> Product Max Rate must be valid (not 0/-)
    </div>
    <div class="rule-box">
    <strong>Rule 3:</strong> If N. Applications is blank/zero/non-numeric = 1
    </div>
    <div class="rule-box filter">
    <strong>Rule 4:</strong> Product Rate UoM must be valid
    </div>
    <div class="rule-box">
    <strong>Rule 5:</strong> Convert Rate to g/ha (Reliance #43988)
    </div>
    <div class="rule-box filter">
    <strong>Rule 6:</strong> Reject if Rate > 20,000 g/ha
    </div>
    <div class="rule-box">
    <strong>Rule 7:</strong> Season Total = Rate(g/ha) x N. Applications
    </div>
    <div class="rule-box filter">
    <strong>Rule 8:</strong> Reject if Season Total >= 1,000,000 g/ha
    </div>
    <div class="rule-box">
    <strong>Rule 9:</strong> Map Target Crop (Reliance #43922)
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title" style="margin-top:20px;">Unit Conversions</div>',
                unsafe_allow_html=True)
    st.caption(f"{len(CONVERSION_FACTORS)} units supported (Reliance #43988)")
    with st.expander("View conversion factors"):
        conv_df = pd.DataFrame(
            [(k, v) for k, v in sorted(CONVERSION_FACTORS.items())],
            columns=["Unit", "Factor to g/ha"]
        )
        st.dataframe(conv_df, use_container_width=True, height=250)

    st.markdown('<div class="section-title" style="margin-top:20px;">About</div>',
                unsafe_allow_html=True)
    st.caption("RIRA Parameters Generator v4.2")
    st.caption("Rules are fixed and cannot be modified by users.")


# ==============================================================================
#  MAIN CONTENT
# ==============================================================================

tab_upload, tab_recommended, tab_alternative = st.tabs(
    ["Upload & Process", "Recommended RIRA Parameters", "Alternative RIRA Parameters per Country of Destination"])

# -- UPLOAD TAB ----------------------------------------------------------------
with tab_upload:
    st.markdown('<div class="section-title">Load GAP Data</div>',
                unsafe_allow_html=True)

    # Option 1: Fetch directly from Qlik
    st.markdown("**Option 1: Fetch from Qlik Cloud (real-time)**")

    force_reload = st.checkbox("Force data reload before fetching (takes 1-5 min)",
                               value=False)              # ← NEW (before the button)

    if st.button("Fetch Data from Qlik", use_container_width=True, type="primary"):
        with st.spinner("Getting Qlik token..."):
            token = get_qlik_token()
        if token:
            if force_reload:
                with st.spinner("Reloading app data from source..."):
                    reload_ok = trigger_qlik_reload(token)
                    if reload_ok:
                        st.success("App reloaded successfully.")
                    else:
                        st.warning("Reload failed or timed out. Using cached data.")

            progress = st.progress(0, text="Connecting to Qlik Cloud...")
            df_qlik, error = fetch_qlik_sync(token, progress)
            progress.empty()
            if error:
                st.error(f"Qlik error: {error}")
            elif df_qlik is not None:
                st.session_state["df_raw"] = df_qlik
                reload_time = df_qlik.attrs.get("last_reload_time", "Unknown")
                st.success(f"Loaded **{len(df_qlik):,}** rows from Qlik Cloud ({len(df_qlik.columns)} columns)")
                st.info(f"📅 Data as of last Qlik reload: **{reload_time}**")
                st.rerun()
        else:
            st.warning("Qlik credentials not configured in secrets.toml")



    st.markdown("---")

    # Option 2: Upload file
    st.markdown("**Option 2: Upload exported file (Excel or CSV)**")

    uploaded_file = st.file_uploader(
        "Drag and drop or browse",
        type=["xlsx", "xls", "csv"],
        help="Export from Qlik Cloud: GAP DATA NON-SEED CARE sheet"
    )

    if uploaded_file is not None:
        with st.spinner("Reading file..."):
            try:
                if uploaded_file.name.endswith(".csv"):
                    df_raw = pd.read_csv(uploaded_file)
                else:
                    df_raw = pd.read_excel(uploaded_file)
                st.session_state["df_raw"] = df_raw
                st.success(f"File loaded: **{uploaded_file.name}** | "
                           f"**{len(df_raw):,}** rows | **{len(df_raw.columns)}** columns")
            except Exception as e:
                st.error(f"Error reading file: {e}")

    if st.session_state.get("df_raw") is not None:
        st.markdown("---")
        if st.button("Process Data (Apply Business Rules)", use_container_width=True,            
                     type="primary"):            
            with st.spinner("Applying business rules..."):
                df_raw = st.session_state["df_raw"]
                processed_df, stats = process_gap_data(df_raw)

                if processed_df is not None:
                    st.session_state["processed_df"] = processed_df
                    st.session_state["stats"] = stats
                    st.success(
                        f"Done! **{stats['final_records']:,}** valid records | "
                        f"**{stats.get('gap_variants', 'N/A')}** GAP Variants | "
                        f"**{stats.get('countries', 'N/A')}** Countries"
                    )
                    st.rerun()

        

        # Raw preview
        st.markdown('<div class="section-title" style="margin-top:20px;">Raw Data Preview</div>',
                    unsafe_allow_html=True)
        st.dataframe(st.session_state["df_raw"].head(50), use_container_width=True, height=250)
        st.caption(f"Showing first 50 of {len(st.session_state['df_raw']):,} rows")

# -- RECOMMENDED RIRA PARAMETERS TAB -------------------------------------------
with tab_recommended:
    st.markdown('<div class="section-title">Recommended RIRA Parameters</div>',
                unsafe_allow_html=True)
    st.caption("One line per GAP Variant. Parameters represent the worst-case scenario "
               "across ALL countries of registration (global maximum).")

    processed_df = st.session_state.get("processed_df")
    stats = st.session_state.get("stats")

    if processed_df is not None:
        # Calculate aggregated RIRA parameters per GAP Variant (all countries)
        def calc_recommended(group):
            # Highest Single Application Rate = MAX rate per application across all records
            highest_single = group["rate_gha"].max()

            # Find the record with the highest season total (rate x n_apps)
            max_season_idx = group["season_total_rate"].idxmax()
            # Maximum Season Application Rate = the RATE PER APPLICATION of that record
            max_season_rate = group.loc[max_season_idx, "rate_gha"]
            # Number of Applications = n_apps of that same record
            max_season_apps = group.loc[max_season_idx, "max_app_numeric"]

            # Find the record with the highest single rate
            max_single_idx = group["rate_gha"].idxmax()

            # RIRA validation: Highest Single must be >= Max Season Rate
            if highest_single < max_season_rate:
                max_season_rate = highest_single

            # Country for Max Season Application Rate
            if "country" in group.columns:
                country_season = group.loc[max_season_idx, "country"]
                country_single = group.loc[max_single_idx, "country"]

                # If same country drives both parameters, show once
                if country_season == country_single:
                    countries_str = str(country_season) if pd.notna(country_season) else ""
                else:
                    parts = []
                    if pd.notna(country_season):
                        parts.append(f"{country_season} (Max Season)")
                    if pd.notna(country_single):
                        parts.append(f"{country_single} (Highest Single)")
                    countries_str = ", ".join(parts)
            else:
                countries_str = ""

            # All crop equivalents registered for this variant
            if "crop_equivalent" in group.columns:
                crops = sorted(group["crop_equivalent"].dropna().unique().tolist())
                crops_str = ", ".join(crops)
            elif "target_crop" in group.columns:
                crops = sorted(group["target_crop"].dropna().unique().tolist())
                crops_str = ", ".join(crops)
            else:
                crops_str = ""

            # Application methods from driving records
            method_list = []
            if "app_method" in group.columns:
                method_season = group.loc[max_season_idx, "app_method"]
                method_single = group.loc[max_single_idx, "app_method"]
                if pd.notna(method_season):
                    method_list.append(str(method_season))
                if pd.notna(method_single) and method_single != method_season:
                    method_list.append(str(method_single))
            methods_str = ", ".join(sorted(set(method_list)))

            return pd.Series({
                "Country": countries_str,
                "Crop Equivalents": crops_str,
                "Application Method": methods_str,
                "Maximum Season Application Rate (g/ha)": round(max_season_rate, 4),
                "Highest Single Application Rate (g/ha)": round(highest_single, 4),
                "Number of Applications": int(max_season_apps),
            })

        recommended_df = processed_df.groupby("gap_variant").apply(calc_recommended).reset_index()
        recommended_df = recommended_df.rename(columns={"gap_variant": "GAP Variant"})

        # Metrics
        c1, c2, c3 = st.columns(3)
        c1.markdown(f'<div class="metric-card"><div class="metric-value">{len(recommended_df)}</div>'
                    f'<div class="metric-label">GAP Variants</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card"><div class="metric-value">{stats["final_records"]:,}</div>'
                    f'<div class="metric-label">Valid Records Processed</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="metric-card"><div class="metric-value">{stats.get("countries", "N/A")}</div>'
                    f'<div class="metric-label">Countries in Dataset</div></div>', unsafe_allow_html=True)

        # Search
        st.markdown("<br>", unsafe_allow_html=True)
        search = st.text_input("Search GAP Variant", "", placeholder="Type to filter...",
                               key="recommended_search")
        display_rec = recommended_df
        if search:
            display_rec = recommended_df[
                recommended_df["GAP Variant"].str.contains(search, case=False, na=False)
            ]

        st.dataframe(display_rec, use_container_width=True, height=500)
        st.caption(f"Showing {len(display_rec)} of {len(recommended_df)} GAP Variants")

        # Download
        st.markdown("<br>", unsafe_allow_html=True)
        buf_rec = BytesIO()
        with pd.ExcelWriter(buf_rec, engine="openpyxl") as writer:
            display_rec.to_excel(writer, index=False, sheet_name="Recommended RIRA Parameters")
        buf_rec.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            "Download Recommended RIRA Parameters (Excel)",
            data=buf_rec,
            file_name=f"Recommended_RIRA_Parameters_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("Upload data and click **Process Data** in the Upload tab.")


# -- ALTERNATIVE RIRA PARAMETERS PER COUNTRY -----------------------------------
with tab_alternative:
    st.markdown('<div class="section-title">Alternative RIRA Parameters per Country of Destination</div>',
                unsafe_allow_html=True)
    st.caption("Parameters calculated per GAP Variant AND Country. "
               "Use this to select country-specific parameters when the global recommendation is not applicable.")

    processed_df = st.session_state.get("processed_df")

    if processed_df is not None and "country" in processed_df.columns:
        def calc_per_country(group):
            highest_single = group["rate_gha"].max()
            max_idx = group["season_total_rate"].idxmax()
            max_season_rate = group.loc[max_idx, "rate_gha"]
            max_season_apps = group.loc[max_idx, "max_app_numeric"]

            if highest_single < max_season_rate:
                max_season_rate = highest_single

            if "crop_equivalent" in group.columns:
                crops = sorted(group["crop_equivalent"].dropna().unique().tolist())
                crops_str = ", ".join(crops)
            elif "target_crop" in group.columns:
                crops = sorted(group["target_crop"].dropna().unique().tolist())
                crops_str = ", ".join(crops)
            else:
                crops_str = ""

            if "app_method" in group.columns:
                methods = sorted(group["app_method"].dropna().unique().tolist())
                methods_str = ", ".join(methods)
            else:
                methods_str = ""

            return pd.Series({
                "Crop Equivalents": crops_str,
                "Application Method": methods_str,
                "Maximum Season Application Rate (g/ha)": round(max_season_rate, 4),
                "Highest Single Application Rate (g/ha)": round(highest_single, 4),
                "Number of Applications": int(max_season_apps),
                "Season Total (g/ha)": round(max_season_rate * max_season_apps, 4),
            })

        alt_df = processed_df.groupby(["gap_variant", "country"]).apply(calc_per_country).reset_index()
        alt_df = alt_df.rename(columns={"gap_variant": "GAP Variant", "country": "Country"})
        alt_df = alt_df.sort_values(["GAP Variant", "Country"]).reset_index(drop=True)

        # Metrics
        c1, c2 = st.columns(2)
        c1.markdown(f'<div class="metric-card"><div class="metric-value">{len(alt_df):,}</div>'
                    f'<div class="metric-label">Variant-Country Combinations</div></div>',
                    unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card"><div class="metric-value">'
                    f'{alt_df["Country"].nunique()}</div>'
                    f'<div class="metric-label">Countries</div></div>', unsafe_allow_html=True)

        # ==================================================================
        # FILTERS - GAP Variant and Country
        # ==================================================================
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">Filters</div>', unsafe_allow_html=True)

        col_filter1, col_filter2 = st.columns(2)

        with col_filter1:
            all_variants = sorted(alt_df["GAP Variant"].unique().tolist())
            selected_variants = st.multiselect(
                "GAP Variant (Design Code)",
                options=all_variants,
                default=[],
                placeholder="All variants (type to search...)",
                key="alt_variant_filter"
            )

        with col_filter2:
            all_countries = sorted(alt_df["Country"].unique().tolist())
            selected_countries = st.multiselect(
                "Country of Destination",
                options=all_countries,
                default=[],
                placeholder="All countries (type to search...)",
                key="alt_country_filter"
            )

        # Apply filters
        display_alt = alt_df.copy()
        if selected_variants:
            display_alt = display_alt[display_alt["GAP Variant"].isin(selected_variants)]
        if selected_countries:
            display_alt = display_alt[display_alt["Country"].isin(selected_countries)]

        # ==================================================================
        # Display filtered results
        # ==================================================================
        st.markdown("<br>", unsafe_allow_html=True)
        st.dataframe(display_alt, use_container_width=True, height=500)
        st.caption(f"Showing {len(display_alt):,} of {len(alt_df):,} rows"
                   f"{' (filtered)' if selected_variants or selected_countries else ''}")

        # Download (exports only the filtered view)
        st.markdown("<br>", unsafe_allow_html=True)
        buf_alt = BytesIO()
        with pd.ExcelWriter(buf_alt, engine="openpyxl") as writer:
            display_alt.to_excel(writer, index=False,
                           sheet_name="Alt RIRA Params per Country")
        buf_alt.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            "Download Alternative RIRA Parameters (Excel)",
            data=buf_alt,
            file_name=f"Alternative_RIRA_Parameters_per_Country_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    elif processed_df is not None:
        st.warning("Country column not found in the data.")
    else:
        st.info("Upload data and click **Process Data** in the Upload tab.")



# -- FOOTER --------------------------------------------------------------------
st.markdown(f"""
<hr style="border:1px solid #E0E0E0; margin-top:40px;">
<p style="text-align:center; color:{SYNGENTA_GRAY}; font-size:12px; font-family:Arial;">
  Syngenta P&S - DPI - RIRA Parameters Generator v4.2 - {datetime.now().year}
</p>
""", unsafe_allow_html=True)